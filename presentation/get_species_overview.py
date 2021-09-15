#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = "Max Hahn-Klimroth"
__copyright__ = "Copyright 2020, M. Hahn-Klimroth, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes"]
__license__ = "MIT"
__version__ = "1.0"
__status__ = "Development"

"""
This script allows to merge the important key values per individual into a joint document
describing the whole species

get_species_overview()
draw_only_timelines(path_to_xlsx, mode = 'total', normalised = False)
"""

import os, io
import numpy as np
from openpyxl import load_workbook, Workbook
from scipy import stats
import copy
import matplotlib.pyplot as plt
from matplotlib.pyplot import *
import pandas as pd
import string


INPUT_FOLDER_KI_AUSWERTUNG = ''
LIST_OF_INDIVIDUALS = []

SPECIES = 'Elen'

DO_TOTAL = False
DO_BINARY = True


OUTPUT_XLSX_TOTAL = ''
OUTPUT_XLSX_BINARY = ''
TIMELINE_OUTPUT_FOLDER = ''
TIMELINE_TITLE = 'Activity budget'

CSV_INDIVIDUAL = 'E:/Jenny/2021-07-21_KI_Auswertung_Elen/Übersicht_Statistik/Individual_Info.csv'


BEHAVIOR_NAMES = {0: 'Standing', 1: 'LHU', 2: 'LHD', 3: 'Out'}
BEHAVIOR_NAMES_BINARY = {0: 'Standing', 1: 'Lying', 3: 'Out'}

COLORS_TO_DRAW = {0:'#104e8b', 1: 'forestgreen', 2:'limegreen', 3:'grey'}
COLORS_TO_DRAW_BINARY = {0:'#104e8b', 1: 'forestgreen', 3:'grey'}

COLORS_TO_DRAW_NORM = {0:'#104e8b', 1: 'forestgreen', 2:'limegreen'}
COLORS_TO_DRAW_BINARY_NORM = {0:'#104e8b', 1: 'forestgreen'}





##############################################################################

HOURS = 14
START_TIME = 17

TIMES = [str(x%24).zfill(2) for x in range(START_TIME, START_TIME+1+HOURS )]
CORRESPONDING_INTERVALS = [np.floor( 3600/7 * i ) for i in range(len(TIMES)) ]


BEHAVIOR_MAPPING = {0:0, 1:1, 2:2, 3:3, 4:3}


def get_alphabet():

    alphabet = list(string.ascii_uppercase)
    ret = []
    drei = []
    ret2 = []
    
    for a in alphabet:
        for b in alphabet:
            ret.append(a+b)
    
    for a in alphabet:
        for b in alphabet:
            for c in alphabet:
                drei.append(a+b+c)
    
    for a in alphabet:
        ret2.append(a)
    for b in ret:
        ret2.append(b)
    for c in drei:
        ret2.append(c)
    
    return ret2
            
            
    
def read_worksheet(wb, wsname):
    alphabet = get_alphabet()
    
    sheet = wb[wsname]
    rows = sheet.rows
    first_row = [cell.value for cell in next(rows)]
    data = {}
    z = 2
    for row in rows:
        record = {}
        j = 0
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[alphabet[j]] = cell.value.strip()
            else:
                record[alphabet[j]] = cell.value
            j += 1
        data[z] = record
        z += 1
    return data






def running_mean(x, N):
    cumsum = np.cumsum(np.insert(x, 0, 0)) 
    return (cumsum[N:] - cumsum[:-N]) / float(N)

def read_single_xlsx(path_to_xlsx):
    
    def _read_sequence_xlsx(wb, behav_map = BEHAVIOR_MAPPING):
        ws_data = read_worksheet(wb, 'Zeitintervalle')
        ret = []
        
        j = 2
        while j <= max( ws_data.keys()) :
            ret.append( behav_map[int(ws_data[j]['F'])] )
            j += 1
        
        return ret
    
    def _read_phases_xlsx(wb):
        ws = read_worksheet(wb, 'Aktivitätsphasen_Übersicht')
        
        standing, lying, sleeping, out = [], [], [], []
        
        for zeile in ws.keys():
            if ws[zeile]['A']:
                standing.append(ws[zeile]['A'])
            if ws[zeile]['B']:
                lying.append(ws[zeile]['B'])
            if ws[zeile]['C']:
                sleeping.append(ws[zeile]['C'])
            if ws[zeile]['D']:
                out.append(ws[zeile]['D'])
        
        return {0: standing, 1: lying, 2: sleeping, 3: out}
    
    sequence = []
    anteil = {0: 0, 1:0, 2:0, 3:0}
    anzahl = {0: 0, 1:0, 2:0, 3:0}
    phasen = {0: [], 1:[], 2:[], 3:[]}
    
    if not os.path.exists(path_to_xlsx):
        print("File not found:", path_to_xlsx)
        return
    
    with open(path_to_xlsx, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = load_workbook(in_mem_file)  

    ws_stat = read_worksheet(wb, 'Statistik')
    
    anzahl = { 0 : int(ws_stat[2]['B']), 
              1: int(ws_stat[2]['C']),
              2: int(ws_stat[2]['D']),
              3: int(ws_stat[2]['E']) }
    
    anteil = { 0 : np.round(float(ws_stat[4]['B'])/100, 4), 
              1: np.round(float(ws_stat[4]['C'])/100, 4), 
              2: np.round(float(ws_stat[4]['D'])/100, 4), 
              3: np.round(float(ws_stat[4]['E'])/100, 4) }
    
    sequence = _read_sequence_xlsx(wb)
    
    phasen = _read_phases_xlsx(wb)
    
    return anzahl, anteil, sequence, phasen


def _read_all_individuals(folder_to_species, species = '', mode = 'total', 
                          individuals_to_process = LIST_OF_INDIVIDUALS):
    
    """ 
    Returns {individual_code: { datum: { anzahl: {0: x, ..., 3:x}, 
                                        anteil: {0: x, ..., 3:x}, 
                                        sequence: [1,2,1,1,2,1,...], 
                                        phasen: {0:[],...,3:[]} } }}
    """
    
    ret = {}
    
    if not os.path.exists(folder_to_species):
        print("Error: Path not found.", folder_to_species)
        return
    if not mode in ['binary', 'total']:
        print("Error: Unknown mode.", mode)
        return
    for zoo in sorted(os.listdir(folder_to_species)):
        curr_path = folder_to_species + zoo + '/'
        for individual in sorted(os.listdir(curr_path)):
            individual_code = '{}_{}_{}'.format(species, zoo, individual)

            if not individual_code in individuals_to_process:
                continue
            curr_path = folder_to_species + zoo + '/' + individual + '/'    
            
            if not os.path.exists(curr_path + mode + '/final/'):
                print("ERROR: Path not found.", curr_path + mode + '/final/')
                continue
            
            ret[individual_code] = {}
            
            files = sorted(os.listdir(curr_path + mode + '/final/'))
            for file in files:
                if not file.endswith('.xlsx'):
                    continue
                datum = file.split('_')[0]
                print('Currently processing:', file)
                anzahl, anteil, sequence, phasen = read_single_xlsx(curr_path + mode + '/final/' + file)
                
                ret[individual_code][datum] = {
                    'anzahl': anzahl,
                    'anteil': anteil,
                    'sequence': sequence,
                    'phasen': phasen
                    }
    
    return ret
                


def write_ws_sequences(wb, inp_dict, ext = '', behavs = BEHAVIOR_NAMES, normalise = False):
    """ 
    inp = {individual_code: { datum: { anzahl: {0: x, ..., 3:x}, 
                                        anteil: {0: x, ..., 3:x}, 
                                        sequence: [1,2,1,1,2,1,...], 
                                        phasen: {0:[],...,3:[]} } }}
    """
    
    inp = copy.deepcopy(inp_dict)
    def _calculate_avg_per_species(inp_dict):
        """ 
        inp_dict = {individual_code: [ {0:0.32, 1:0.56, 2:0....} ] of length 7200 }
        """
        def _get_entries(input_dict, coordinate, behavior, normalise = normalise):

            ret = []
            for ind in inp_dict:
                if normalise:
                    other = 0
                    for b in inp_dict[ind][coordinate].keys():
                        if b != 3:
                            other += inp_dict[ind][coordinate][b]
                    if other > 0 and inp_dict[ind][coordinate][3] > 0:
                        for b in inp_dict[ind][coordinate].keys():
                            if b != 3:
                                inp_dict[ind][coordinate][b] = inp_dict[ind][coordinate][b]/other
                            inp_dict[ind][coordinate][3] = 0
                                
                ret.append( inp_dict[ind][coordinate][behavior] )
            return ret
        
        avg = {}
        sem = {}      
        tmp = {}

        for i in behavs.keys():
            tmp[i] = []
            for j in range(7200):
                x = _get_entries(input_dict = inp_dict, coordinate = j, behavior = i)
                tmp[i].append(x)
        
        
        if normalise:
            other = 0
            for j in range(7200):
                other = 0
                for b in tmp.keys():
                    if b != 3:
                        other += sum(tmp[b][j])
                if other > 0 and sum(tmp[3][j]) > 0:
                    
                    out_pos = []
                    for k in range(len( tmp[3][j] )):
                        if tmp[3][j][k] > 0:
                            out_pos.append(k)
                    

                    for k in out_pos:
                        for b in tmp.keys():
                            tmp[b][j].pop(k)

                    

                        
        
        for i in behavs.keys():
            
            avg[i] = []
            sem[i] = []
                
            
            for j in range(7200):                
                
                
                avg[i].append( np.mean(tmp[i][j]) )
                if len(tmp[i][j]) > 1:
                    sem[i].append( stats.sem(tmp[i][j]) )
                else:
                    sem[i].append(0.0)
        
        
        
        
        return avg, sem
            
         
                        
        
            
            
    
    def _calculate_avg_per_individual(inp_dict, normalise = normalise):
        """
        datum: { anzahl: {0: x, ..., 3:x}, 
                                        anteil: {0: x, ..., 3:x}, 
                                        sequence: [1,2,1,1,2,1,...], 
                                        phasen: {0:[],...,3:[]} } }
        """

        
        
        seq = [{0:0, 1:0, 2:0, 3:0} for j in range(7200)]

        nights = 0
        for date in inp_dict.keys():
            nights += 1
            curr_seq = inp_dict[date]['sequence']
            for j in range(min(len(curr_seq), 7200)):
                seq[j][curr_seq[j]] += 1

        avg = [{0:0, 1:0, 2:0, 3:0} for j in range(7200)]
        
        for j in range(len(seq)):
            for i in behavs.keys():
                avg[j][i] = np.float( seq[j][i] / nights )
        
        
        if normalise:
            for j in range(len(seq)):
                other = 0
                for i in behavs.keys():
                    if i == 3:
                        continue
                    other += avg[j][i]
                if other > 0 and avg[j][3] > 0:
                    for i in behavs.keys():
                        avg[j][i] = avg[j][i] / other
                    avg[j][3] = 0
        
        return avg        
        
        
    
    ws = wb.create_sheet('Sequences'+ext)
    ws_avg_ind = wb.create_sheet('Seq-MEAN-IND'+ext)
    
    ws_avg = wb.create_sheet('Seq-MEAN'+ext)
    ws_sem = wb.create_sheet('Seq-SEM'+ext)
    
    
    ws.cell(1,1).value = 'Individual'
    ws.cell(2,1).value = 'Date'
    
    ws_avg_ind.cell(1,1).value = 'Individual'
    ws_avg_ind.cell(2,1).value = 'Behaviour'
    
    ws_avg.cell(1,1).value = 'Interval'
    ws_sem.cell(1,1).value = 'Interval'
    
    spalte = 2
    
    avgs = {}
    
    for i in range(7201):
        ws.cell(3+i, 1).value = i+1
        ws_avg_ind.cell(3+i, 1).value = i + 1 
        ws_avg.cell(2+i, 1).value = i + 1 
        ws_sem.cell(2+i, 1).value = i + 1 
    
    avg_spalte = 2
    for ind in sorted(inp.keys()):
        
        for date in sorted(inp[ind].keys()):
            ws.cell(1,spalte).value = ind
            ws.cell(2,spalte).value = date
            curr_seq = inp[ind][date]['sequence']
            for j in range(len(curr_seq)):
                ws.cell(3+j, spalte).value = curr_seq[j]
            spalte += 1
        
        avg = _calculate_avg_per_individual(inp_dict = inp[ind])
        
        x = 0
        for i in sorted(behavs.keys()):
            ws_avg_ind.cell(1, avg_spalte+x).value = ind
            ws_avg_ind.cell(2, avg_spalte+x).value = behavs[i]
            for j in range(len(avg)):
                ws_avg_ind.cell(3+j, avg_spalte+x).value = avg[j][i]
            x += 1
        avg_spalte += len(behavs.keys())
        
        avgs[ind] = avg
    
    total_avg, total_sem = _calculate_avg_per_species(avgs)
    spalte = 2
    x = 0
    for i in sorted(behavs.keys()):
        ws_avg.cell(1, spalte+x).value = behavs[i]
        ws_sem.cell(1, spalte + x).value = behavs[i]
        for j in range(len(total_avg[i])):
            ws_avg.cell(2+j, spalte + x).value = total_avg[i][j]
            ws_sem.cell(2+j, spalte + x).value = total_sem[i][j]
        x += 1




def normalise_data(inp_dict):
    """
    Scales the percentage of each behavior in accordance to being out.
    """

    ret = copy.deepcopy(inp_dict)
    for ind in inp_dict.keys():
        for datum in inp_dict[ind].keys():
            if inp_dict[ind][datum]['anteil'][3] > 0:
                other = inp_dict[ind][datum]['anteil'][0] + inp_dict[ind][datum]['anteil'][1] + inp_dict[ind][datum]['anteil'][2]
                if other > 0:
                    ret[ind][datum]['anteil'][0] = inp_dict[ind][datum]['anteil'][0] / other
                    ret[ind][datum]['anteil'][1] = inp_dict[ind][datum]['anteil'][1] / other
                    ret[ind][datum]['anteil'][2] = inp_dict[ind][datum]['anteil'][2] / other
                    ret[ind][datum]['anteil'][3] = 0

    return ret

        
    
def write_overview(wb, inp_dict, ext = '', behavs = BEHAVIOR_NAMES):
    """ {individual_code: { datum: { anzahl: {0: x, ..., 3:x}, 
                                        anteil: {0: x, ..., 3:x}, 
                                        sequence: [1,2,1,1,2,1,...], 
                                        phasen: {0:[],...,3:[]} } }} """
    
    ws_overview = wb.create_sheet('Overview')
    ws_overview_avg = wb.create_sheet('Overview-Mean')
    
    ws_overview.cell(1,1).value = 'Individual'
    ws_overview.cell(1,2).value = 'Date'
    ws_overview.cell(1,3).value = 'Behaviour'
    ws_overview.cell(1,4).value = 'Amount'
    ws_overview.cell(1,5).value = 'Count'
    ws_overview.cell(1,6).value = 'Cod_short'
    ws_overview.cell(1,7).value = 'Cod_ssn'
    ws_overview.cell(1,8).value = 'Species'
    ws_overview.cell(1,9).value = 'Age'
    ws_overview.cell(1,10).value = 'Sex'
    ws_overview.cell(1,11).value = 'Zoo'
    ws_overview.cell(1,12).value = 'Stabeling'
    ws_overview.cell(1,13).value = 'Stable'
    ws_overview.cell(1,14).value = 'Cat'
    
    
    ws_overview_avg.cell(1,1).value = 'Individual'
    ws_overview_avg.cell(1,2).value = 'Behaviour'
    ws_overview_avg.cell(1,3).value = 'Amount_AVG'
    ws_overview_avg.cell(1,4).value = 'Count_AVG'
    ws_overview_avg.cell(1,5).value = 'Amount_SEM'
    ws_overview_avg.cell(1,6).value = 'Count_SEM'
    ws_overview_avg.cell(1,7).value = 'median_phase_len'
    ws_overview_avg.cell(1,8).value = 'first_quantile_phase_len'
    ws_overview_avg.cell(1,9).value = 'third_quantile_phase_len'    
    ws_overview_avg.cell(1,10).value = 'Cod_short'
    ws_overview_avg.cell(1,11).value = 'Cod_ssn'
    ws_overview_avg.cell(1,12).value = 'Species'
    ws_overview_avg.cell(1,13).value = 'Age'
    ws_overview_avg.cell(1,14).value = 'Sex'
    ws_overview_avg.cell(1,15).value = 'Zoo'
    ws_overview_avg.cell(1,16).value = 'Stabeling'
    ws_overview_avg.cell(1,17).value = 'Stable'
    ws_overview_avg.cell(1,18).value = 'Cat'
    
    
    
    
    
    
    anz_avgs = {}
    ant_avgs = {}
    anz_sems = {}
    ant_sems = {}
    phases = {}
    
    add_info = get_additional_information_per_ind()
    
    zeile = 2
    zeile_avg = 2
    

    
    for ind in sorted(inp_dict.keys()):
        anzahl_ind = {}
        anteil_ind = {}
        anz_avgs[ind] = {}
        ant_avgs[ind] = {}
        ant_sems[ind] = {}
        anz_sems[ind] = {}
        
        phases[ind] = {}
        
        for i in sorted(behavs.keys()): 
            anzahl_ind[i] = []
            anteil_ind[i] = []
            phases[ind][i] = []
            
           
        for datum in sorted(inp_dict[ind].keys()):
            anz = inp_dict[ind][datum]['anzahl']
            anteil = inp_dict[ind][datum]['anteil']
            
            for i in sorted(behavs.keys()):                
                
                anzahl_ind[i].append(anz[i])
                anteil_ind[i].append(anteil[i])
                phases[ind][i] += inp_dict[ind][datum]['phasen'][i]
                
                ws_overview.cell(zeile, 1).value = ind
                ws_overview.cell(zeile, 2).value = datum
                ws_overview.cell(zeile, 3).value = behavs[i]
                
                ws_overview.cell(zeile, 4).value = anteil[i]
                ws_overview.cell(zeile, 5).value = anz[i]
                
                ws_overview.cell(zeile,6).value = add_info[ind]['Cod_short']
                ws_overview.cell(zeile,7).value = add_info[ind]['Cod_ssn']
                ws_overview.cell(zeile,8).value = add_info[ind]['Species']
                ws_overview.cell(zeile,9).value = add_info[ind]['Age']
                ws_overview.cell(zeile,10).value = add_info[ind]['Sex']
                ws_overview.cell(zeile,11).value = add_info[ind]['Zoo']
                ws_overview.cell(zeile,12).value = add_info[ind]['Stabeling']
                ws_overview.cell(zeile,13).value = add_info[ind]['Stable']
                ws_overview.cell(zeile,14).value = add_info[ind]['Cat']
                
                
                zeile += 1
        
        for i in sorted(behavs.keys()): 
            anz_avgs[ind][i] = np.mean(anzahl_ind[i]) 
            ant_avgs[ind][i] = np.mean(anteil_ind[i]) 
            ant_sems[ind][i] = stats.sem( anteil_ind[i] )
            anz_sems[ind][i] = stats.sem( anzahl_ind[i] )
            
            
        
            ws_overview_avg.cell(zeile_avg,1).value = ind
            ws_overview_avg.cell(zeile_avg,2).value = behavs[i]
            ws_overview_avg.cell(zeile_avg,3).value = ant_avgs[ind][i]
            ws_overview_avg.cell(zeile_avg,4).value = anz_avgs[ind][i]
            ws_overview_avg.cell(zeile_avg,5).value = ant_sems[ind][i]
            ws_overview_avg.cell(zeile_avg,6).value = anz_sems[ind][i]
            
            ws_overview_avg.cell(zeile_avg,10).value = add_info[ind]['Cod_short']
            ws_overview_avg.cell(zeile_avg,11).value = add_info[ind]['Cod_ssn']
            ws_overview_avg.cell(zeile_avg,12).value = add_info[ind]['Species']
            ws_overview_avg.cell(zeile_avg,13).value = add_info[ind]['Age']
            ws_overview_avg.cell(zeile_avg,14).value = add_info[ind]['Sex']
            ws_overview_avg.cell(zeile_avg,15).value = add_info[ind]['Zoo']
            ws_overview_avg.cell(zeile_avg,16).value = add_info[ind]['Stabeling']
            ws_overview_avg.cell(zeile_avg,17).value = add_info[ind]['Stable']
            ws_overview_avg.cell(zeile_avg,18).value = add_info[ind]['Cat']
            
            ws_overview_avg.cell(zeile_avg,7).value = np.median( phases[ind][i] )
            try:
                ws_overview_avg.cell(zeile_avg,8).value = np.quantile( phases[ind][i], 0.25 )
                ws_overview_avg.cell(zeile_avg,9).value = np.quantile( phases[ind][i], 0.75 )
            except:
                ws_overview_avg.cell(zeile_avg,8).value = 0
                ws_overview_avg.cell(zeile_avg,9).value = 0
            zeile_avg += 1
        
        
    
    
def ensure_directory(filename):
    folder = os.path.dirname(filename)
    if not os.path.exists(folder):
        os.makedirs(folder)
          



def get_additional_information_per_ind(csv_ind = CSV_INDIVIDUAL, 
                                       #csv_species_text = CSV_SPECIES_TEXT 
                                       #csv_species_short = CSV_SPECIES_SHORT
                                       ):
    """
    Input: Three csv files containing the information that should be printed 
    in each coloumn of the final sheet.
    Output: { individual_code: {'x': value}
             }
    with x ranging in 
        Cod_short	Cod_ssn	Species	Age	Sex	Zoo	Stabeling	Stable   Cat

        
        Species, Species_ssn, Genus, Tribe, Family, Order, 
        Head-body[cm], shoulder_height[cm], weight[kg], feeding_type,
        LR_open, LR_medium, LR_close, LR-1 _Forest, LR-2_Savanna, LR-3_Shrubland,
        LR-4 _Grassland, LR-5_Wetlands,	LR-6_Rocky areas, LR-7_Desert

        
       Genus-Text,Tribe-Text,Family-Text,Order-Text

    """
    
    df_ind = pd.read_csv(csv_ind)
# =============================================================================
#     df_spec_txt = pd.read_csv(csv_species_text)
#     df_spec_short = pd.read_csv(csv_species_short)
# =============================================================================
    
    ret = {}
    for row in df_ind.index:
        individual_code = df_ind['Cod_long'][row]
        ret[individual_code] = {}
        for col in df_ind:
            if col == 'Cod_long':
                continue
            ret[individual_code][col] = df_ind[col][row]
    
# =============================================================================
#     spec_info_txt = df_spec_txt[ df_spec_txt['Species'] == species_of_interest]
#     spec_info_short = df_spec_txt[ df_spec_short['Species'] == species_of_interest]
# 
#     if len(spec_info_txt.index) == 0:
#         return {}
#     
#     for col in spec_info_txt:
#         if col == 'Species':
#             continue
#         for individual_code in ret.keys():
#             ret[individual_code][col] = spec_info_txt.iloc[0][col]
#     
#     for individual_code in ret.keys():
#         ret[individual_code]['Genus-Text'] = spec_info_short.iloc[0]['Genus']
#         ret[individual_code]['Tribe-Text'] = spec_info_short.iloc[0]['Tribe']
#         ret[individual_code]['Family-Text'] = spec_info_short.iloc[0]['Family']
#         ret[individual_code]['Order-Text'] = spec_info_short.iloc[0]['Order']
# =============================================================================
    
    return ret
    
        
    
    
    



def create_timeline(path_to_xlsx, 
                    behaviors = BEHAVIOR_NAMES,
                    draw = COLORS_TO_DRAW,
                    ext = '', 
                    timeline_output = TIMELINE_OUTPUT_FOLDER,
                    species = SPECIES,
                    title_string = TIMELINE_TITLE,
                    x_labels = TIMES, 
                    x_ticks = CORRESPONDING_INTERVALS,
                    ):
    
    def read_averages(wb, behaviors = behaviors):
        ws = wb['Overview-Mean']
        max_row = ws.max_row
        
        averages = {}
        for b in behaviors.keys():
            averages[b] = []

        for j in range(2, max_row + 1):
            for h in behaviors.keys():
                if ws.cell(j, 2).value == behaviors[h]:
                    averages[h].append( float(ws.cell(j, 3).value) )
        
        return averages
    
    if not os.path.exists(path_to_xlsx):
        print("Error: Path not found.", path_to_xlsx)
    wb = load_workbook(path_to_xlsx)
    ws_avg = wb['Seq-MEAN']
    ws_sem = wb['Seq-SEM']
    
    
    rcParams['mathtext.fontset'] = 'custom'
    rcParams['mathtext.it'] = 'Arial:italic'
    rcParams['mathtext.rm'] = 'Arial'
    
    cm = 1/2.54  # centimeters in inches
    
    plt.figure(figsize=(18*cm,8*cm), dpi = 600)
    
    curves = []
    legend_curves = []
    ensure_directory(timeline_output)
    
    overall_averages = read_averages(wb)
    
    avgs = {}
    sems = {}
        
    for b in draw.keys():
        avgs[b] = []
        sems[b] = []
            
    for zeile in range(2, ws_avg.max_row):
        j = 0
        for b in avgs.keys():
            
            avgs[b].append( ws_avg.cell(zeile, 2 + j).value )
            sems[b].append( ws_sem.cell(zeile, 2 + j).value )
            j+=1
    
    for b in avgs.keys():
        avgs[b] = running_mean(avgs[b], 27)
        sems[b] = running_mean(sems[b], 27)
    
    upper = {}
    lower = {}
    
    for b in avgs.keys():
        upper[b] = np.array([ avgs[b][j] + sems[b][j] for j in range(min(len(avgs[b]), len(sems[b])) -1) ], dtype='float32')
        lower[b] = np.array([ avgs[b][j] - sems[b][j] for j in range(min(len(avgs[b]), len(sems[b])) -1) ], dtype='float32')
    
    

    for h in avgs.keys(): 
        x, = plt.plot(avgs[h][:-1], color=draw[h], linestyle = '-', 
                     linewidth=0.4, label = str(h))
        curves.append(x)
        legend_curves.append(x)
        
        u, = plt.plot(upper[h], color=draw[h], linestyle = ':', alpha = 0.2,
                     linewidth=0.15, label = 'u'+str(h))
        curves.append(u)
        
        l, = plt.plot(lower[h], color=draw[h], linestyle = ':', alpha = 0.2, 
                     linewidth=0.15, label = 'l'+str(h))
        curves.append(l)
        x_coord = np.array([j for j in range(len(avgs[h][:-1]))], dtype='float32')
        
        plt.fill_between(x_coord, lower[h], upper[h], facecolor = draw[h], alpha = 0.2, interpolate = True)
        

        draw_avg = [ np.mean(overall_averages[h]) for j in range(len(avgs[h][:-1]))]
        sem_plus = [ np.mean(overall_averages[h]) + stats.sem(overall_averages[h]) for j in range(len(avgs[h][:-1]))]
        sem_minus = [ np.mean(overall_averages[h]) - stats.sem(overall_averages[h])  for j in range(len(avgs[h][:-1]))]
            
        plt.plot(draw_avg, color=draw[h], linestyle='-', linewidth=0.75)
        plt.plot(sem_plus, color=draw[h], linestyle=':', linewidth=0.75)
        plt.plot(sem_minus, color=draw[h], linestyle=':', linewidth=0.75)
        
    plt.ylabel('Percentage', fontsize=8)
    plt.xlabel('Time', fontsize=8)
    plt.xticks(x_ticks, x_labels, fontsize=6)
    plt.yticks([np.round(x/10,1) for x in range(0, 11) ], fontsize=6)
        
    plt.legend( legend_curves, [behaviors[h] for h in draw.keys()], loc='lower right',
          fancybox=True, shadow=False, fontsize=7, borderaxespad = 0., bbox_to_anchor=(1.0, 1.03), ncol=len(draw.keys()) )
    plt.title(title_string, fontsize=10)
        
    axes = plt.gca()
    axes.margins(x=0.005)
    axes.set_ylim([-0.02, 1.02])
        
    plt.savefig(timeline_output + species + ext + '.jpg')
    plt.close()



def write_phase_overview(ws, overview, behavs):
    """ Input: {individual_code: { datum: { anzahl: {0: x, ..., 3:x}, 
                                        anteil: {0: x, ..., 3:x}, 
                                        sequence: [1,2,1,1,2,1,...], 
                                        phasen: {0:[],...,3:[]} } }}"""
    ws.title = "Phases"
    ws.cell(1,1).value = 'Individual'
    ws.cell(1,2).value = 'Behaviour'
    ws.cell(1,3).value = 'Date'
    ws.cell(1,4).value = 'Duration'
    
    ws.cell(1,5).value = 'Cod_short'
    ws.cell(1,6).value = 'Cod_ssn'
    ws.cell(1,7).value = 'Species'
    ws.cell(1,8).value = 'Age'
    ws.cell(1,9).value = 'Sex'
    ws.cell(1,10).value = 'Zoo'
    ws.cell(1,11).value = 'Stabeling'
    ws.cell(1,12).value = 'Stable'
    ws.cell(1,12).value = 'Cat'
    
    
    
    add_info = get_additional_information_per_ind()
    
    
                   
    
    zeile = 2
    for individual_code in sorted(overview.keys()):
        for date in sorted(overview[individual_code]):
            for b in sorted(overview[individual_code][date]['phasen'].keys()):
                if not b in behavs.keys():
                    continue
                for x in overview[individual_code][date]['phasen'][b]:
                    ws.cell(zeile, 1).value = individual_code
                    ws.cell(zeile, 2).value = behavs[b]
                    ws.cell(zeile, 3).value = date
                    ws.cell(zeile, 4).value = x
                    
                    ind = individual_code
                    ws.cell(zeile,5).value = add_info[ind]['Cod_short']
                    ws.cell(zeile,6).value = add_info[ind]['Cod_ssn']
                    ws.cell(zeile,7).value = add_info[ind]['Species']
                    ws.cell(zeile,8).value = add_info[ind]['Age']
                    ws.cell(zeile,9).value = add_info[ind]['Sex']
                    ws.cell(zeile,10).value = add_info[ind]['Zoo']
                    ws.cell(zeile,11).value = add_info[ind]['Stabeling']
                    ws.cell(zeile,12).value = add_info[ind]['Stable']
                    ws.cell(zeile,12).value = add_info[ind]['Cat']
                    zeile += 1
    
    
def get_species_overview(base_folder = INPUT_FOLDER_KI_AUSWERTUNG,
                         species = SPECIES,
                         individuals = LIST_OF_INDIVIDUALS,
                         output_xlsx_total = OUTPUT_XLSX_TOTAL,
                         output_xlsx_binary = OUTPUT_XLSX_BINARY,
                         behav = BEHAVIOR_NAMES,
                         behav_binary = BEHAVIOR_NAMES_BINARY):
    
    
    if DO_TOTAL:
        print('**************************************************************')
        print('Read files to memory: total')
        print('**************************************************************')
        total_overview = _read_all_individuals(base_folder + species + '/', 
                                               species = species, mode = 'total', 
                                               individuals_to_process = individuals)
        
        
        print('**********************')
        print('Write output: total')
        print('**********************')
        wb_total = Workbook()
        
        write_phase_overview(ws = wb_total.active, overview = total_overview, behavs = behav)
        write_ws_sequences(wb_total, total_overview, ext = '', behavs = behav)
        write_overview(wb=wb_total, inp_dict = total_overview, ext = '', behavs=behav)
        
        ensure_directory(output_xlsx_total)
        wb_total.save(output_xlsx_total)
        
        create_timeline(path_to_xlsx = output_xlsx_total, ext = '',
                        behaviors = BEHAVIOR_NAMES,
                        draw = COLORS_TO_DRAW)
        
        
        print('**********************')
        print('Write output: total - normalised')
        print('**********************')
        wb_totalnorm = Workbook()
        total_norm = normalise_data(total_overview)
        
    
        write_phase_overview(ws = wb_totalnorm.active, overview = total_norm, behavs = behav)
        write_ws_sequences(wb_totalnorm, total_norm, ext = '', behavs = behav, normalise = True)
        write_overview(wb=wb_totalnorm, inp_dict = total_norm, ext = '', behavs=behav)
        
        out_totalnorm = output_xlsx_total[:-5] + '-normalised.xlsx'
        ensure_directory(out_totalnorm)
        wb_totalnorm.save(out_totalnorm)
        
        create_timeline(path_to_xlsx = out_totalnorm, ext = '-normalised',
                        behaviors = BEHAVIOR_NAMES,
                        draw = COLORS_TO_DRAW_NORM)
    
    if DO_BINARY:
        print('**************************************************************')
        print('Read files to memory: binary')
        print('**************************************************************')
        
        binary_overview = _read_all_individuals(base_folder + species + '/', 
                                               species = species, mode = 'binary', 
                                               individuals_to_process = individuals)
        
        binary_norm = normalise_data(binary_overview)
        
        print('**********************')
        print('Write output: binary')
        print('**********************')
        
        wb_binary = Workbook()
        write_phase_overview(ws = wb_binary.active, overview = binary_overview, behavs = behav_binary)
        write_ws_sequences(wb_binary, binary_overview, ext = '', behavs = behav_binary)
        write_overview(wb=wb_binary, inp_dict = binary_overview, ext = '', behavs=behav_binary)
        
        ensure_directory(output_xlsx_binary)
        wb_binary.save(output_xlsx_binary)
        
        create_timeline(path_to_xlsx = output_xlsx_binary, ext = '_binary',
                        behaviors = BEHAVIOR_NAMES_BINARY,
                        draw = COLORS_TO_DRAW_BINARY)
        
        print('**********************')
        print('Write output: binary - normalised')
        print('**********************')
        
        wb_binarynorm = Workbook()
        write_phase_overview(ws = wb_binarynorm.active, overview = binary_norm, behavs = behav_binary)
        write_ws_sequences(wb_binarynorm, binary_norm, ext = '', behavs = behav_binary, normalise = True)
        write_overview(wb=wb_binarynorm, inp_dict = binary_norm, ext = '', behavs=behav_binary)
        
        out_binarynorm = output_xlsx_binary[:-5] + '-normalised.xlsx'
        ensure_directory(out_binarynorm)
        wb_binarynorm.save(out_binarynorm)
        
        create_timeline(path_to_xlsx = out_binarynorm, ext = '_binary-normalised',
                        behaviors = BEHAVIOR_NAMES_BINARY,
                        draw = COLORS_TO_DRAW_BINARY_NORM)
    
def draw_only_timelines(path_to_xlsx, mode = 'total', normalised = False):
    if mode == 'binary':
        if normalised:
            create_timeline(path_to_xlsx = path_to_xlsx, ext = '_binary-normalised',
                    behaviors = BEHAVIOR_NAMES_BINARY,
                    draw = COLORS_TO_DRAW_BINARY_NORM)
        else:
            create_timeline(path_to_xlsx = path_to_xlsx, ext = '_binary',
                    behaviors = BEHAVIOR_NAMES_BINARY,
                    draw = COLORS_TO_DRAW_BINARY)
        
    if mode == 'total':
        if normalised:
            create_timeline(path_to_xlsx = path_to_xlsx, ext = '-normalised',
                    behaviors = BEHAVIOR_NAMES,
                    draw = COLORS_TO_DRAW_NORM)
        else:
            create_timeline(path_to_xlsx = path_to_xlsx, ext = '',
                    behaviors = BEHAVIOR_NAMES,
                    draw = COLORS_TO_DRAW)
    