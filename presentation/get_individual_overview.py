#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = "Max Hahn-Klimroth"
__copyright__ = "Copyright 2021, M. Hahn-Klimroth, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert"]
__license__ = "GPL-3.0"
__version__ = "1.0"
__status__ = "Development"

"""
Use only if the data files report standing, lying, sleeping, out in this order.
- merge_predictions()
- draw_timelines()

- create_timeline_from_xlsx(inputxlsx, outputjpg)
"""


import numpy as np
import os, string
from openpyxl import Workbook, load_workbook
from scipy.stats import sem as SEM
from collections import Counter
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt



INDIVIDUALS_TO_MERGE = [ ]


KI_AUSWERTUNG = 'E:/KI_Auswertung/'
OUTPUT_FOLDER_XLSX = 'E:/Output/'
OUTPUT_FOLDER_TIMELINE = 'E:/Output/'
INDIVIDUAL_INFO_CSV = 'E:/individual_info.csv'


VIDEO_START_TIME = 17
VIDEO_HOUR = 14



CLASSIFICATION_MODE = 'binary' # total, binary
BEHAVIORS = ['Standing', 'LHU', 'LHD', 'Out']

""" *************************************************************  """

if CLASSIFICATION_MODE == 'binary':
    MODE = 'binary/'
    EXT = '_binary'
    DRAW_BEHAVS = {0: '#104e8b', 1: 'forestgreen'}
    BEHAVIOR_MAPPING = {0: 0, 1:1, 2:1, 3:3, 4:3}
else:
    MODE = 'total/'
    EXT = ''
    DRAW_BEHAVS = {0: '#104e8b', 1: 'forestgreen', 2: 'lightgreen'}
    BEHAVIOR_MAPPING = {0:0, 1:1, 2:2, 3:3, 4:3}






def _get_speczoo(l):
    spec, zoo, ind = [], [], []
    for indcode in l:
        s, z, i = indcode.split('_')
        spec.append(s)
        zoo.append(z)
        ind.append(i)
    return spec, zoo, ind


SPECIES, ZOO, INDIVIDUAL = _get_speczoo(INDIVIDUALS_TO_MERGE)

ENCLOSURE = INDIVIDUAL
NUMBER_INDIVIDUALS = len(ENCLOSURE)


INDIVIDUAL_CODE = [SPECIES[j] + '_' + ZOO[j] +'_' + INDIVIDUAL[j] for j in range(NUMBER_INDIVIDUALS)]
ENCLOSURE_CODE = [SPECIES[j] + '/' + ZOO[j] +'/' + ENCLOSURE[j] + '/' for j in range(NUMBER_INDIVIDUALS)]

CONTAINING_FOLDER = [KI_AUSWERTUNG + ENCLOSURE_CODE[j] + MODE + 'final/' for j in range(NUMBER_INDIVIDUALS)]
OUTPUT_XLSX = [OUTPUT_FOLDER_XLSX + SPECIES[j]+'/'+INDIVIDUAL_CODE[j]+'_Übersicht_Stat'+EXT+'.xlsx' for j in range(NUMBER_INDIVIDUALS)]
GRAPHIC_TITLE = [INDIVIDUAL_CODE[j] for j in range(NUMBER_INDIVIDUALS)]


TIMELINE = [OUTPUT_FOLDER_TIMELINE+SPECIES[j]+'/'+INDIVIDUAL_CODE[j]+'_Timeline'+EXT+'.jpg' for j in range(NUMBER_INDIVIDUALS)]

#SPECIFIC_STARTING_TIMES = {
#    'Blessbock_Gelsenkirchen': 18,
#    }
#SPECIFIC_VIDEO_LENGTHS = {
#    'Blessbock_Gelsenkirchen': 12
#    }

TIMES = []
HOURS = []
for indcode in INDIVIDUALS_TO_MERGE:
    spec, zoo, _ = indcode.split('_')
    speczoo = spec + '_' + zoo
#    if speczoo in SPECIFIC_STARTING_TIMES.keys():
#        start = SPECIFIC_STARTING_TIMES[speczoo]
#    else:
#        start = VIDEO_START_TIME
#    if speczoo in SPECIFIC_VIDEO_LENGTHS.keys():
#        hours = SPECIFIC_VIDEO_LENGTHS[speczoo]
#    else:
#        hours = VIDEO_HOUR
    
    start = VIDEO_START_TIME
    hours = VIDEO_HOUR
    TIMES.append( [str(x%24).zfill(2)  for x in range(start, start+1+hours )]  )
    HOURS.append(hours)

CORRESPONDING_INTERVALS = [ [np.floor( 3600/7 * i ) for i in range(len(TIMES[j])) ] for j in range(NUMBER_INDIVIDUALS) ]
CUT_OFF = [int(HOURS[j]*60*60/7) for j in range(NUMBER_INDIVIDUALS)]


def get_individualnames(csv_ind = INDIVIDUAL_INFO_CSV):
    df_ind = pd.read_csv(csv_ind)
    ret = {}
    
    for row in df_ind.index:
        individual_code = df_ind['Cod_long'][row]
        ret[individual_code] = df_ind['Cod_ssn'][row]
        
    return ret

def get_alphabet():

    alphabet = list(string.ascii_uppercase)
    ret = []
    ret2 = []
    
    for a in alphabet:
        for b in alphabet:
            ret.append(a+b)
    
    for a in alphabet:
        ret2.append(a)
    for b in ret:
        ret2.append(b)
    
    return ret2
            
            
    
def read_worksheet(wb, wsname):
    alphabet = get_alphabet()
    
    sheet = wb[wsname]
    rows = sheet.rows
    first_row = [cell.value for cell in next(rows)]
    data = {}
    z = 2
    for row in rows:
        record = {}
        j = 0
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[alphabet[j]] = cell.value.strip()
            else:
                record[alphabet[j]] = cell.value
            j += 1
        data[z] = record
        z += 1
    return data



def ensure_directory(filename):
    path = os.path.dirname(os.path.abspath(filename))
    if not os.path.exists(path):
        os.makedirs(path)



def get_sequence(wb, behav_mapping = BEHAVIOR_MAPPING):
    
    ws_data = read_worksheet(wb, 'Zeitintervalle')
    ret = []
    
    j = 2
    while j <= max( ws_data.keys()) :
        ret.append( behav_mapping[int(ws_data[j]['F'])] )
        j += 1
    
    return ret


def get_statistics(wb):
    
    ws = read_worksheet(wb, 'Statistik')
    
    anzahl = [ int(ws[2]['B']), int(ws[2]['C']), int(ws[2]['D']), int(ws[2]['E']) ]
    dauer = [ np.round(float(ws[3]['B']), 2), np.round(float(ws[3]['C']), 2), 
             np.round(float(ws[3]['D']), 2), np.round(float(ws[3]['E']), 2)]
    return anzahl, dauer
   

def _get_phase_lists(wb):
    
    ws = read_worksheet(wb, 'Aktivitätsphasen_Übersicht')
    alphabet = list(string.ascii_uppercase)
    
    ret = [ [], [], [], [] ] 
    j = 2
    while j <= max(ws.keys()):
        for i in range(1, 5):
            x = ws[j][alphabet[i-1]]
            if x:
                ret[i-1].append(int(x))
        j += 1
    
    return ret


def _xlsx_to_prediction(xlsx_file):
    data_wb =  load_workbook(xlsx_file)
    
    sequence = get_sequence(data_wb)
    anzahl, dauer = get_statistics(data_wb)
    phase_lists = _get_phase_lists(data_wb)
    
    
    data_wb.close()
    
    return anzahl, dauer, sequence, phase_lists

def _get_interval_labels(xlsx):
    ret = [[], [], []]
    data_wb =  load_workbook(xlsx)
    
    ws = read_worksheet(data_wb, 'Zeitintervalle')
    alphabet = list(string.ascii_uppercase)
    

    j = 2
    while j <= max(ws.keys()):
        for i in range(1, 4):
            x = ws[j][alphabet[i-1]]
            ret[i-1].append(x)
        j += 1
    
    return ret
            

def running_mean(x, N):
    cumsum = np.cumsum(np.insert(x, 0, 0)) 
    return (cumsum[N:] - cumsum[:-N]) / float(N)


def create_timeline_from_xlsx(xlsx_file, output_fig, draw = DRAW_BEHAVS,
                              x_labels = TIMES[0], 
                              x_ticks = CORRESPONDING_INTERVALS[0],
                              behavs = BEHAVIORS,
                              title_string = GRAPHIC_TITLE[0],
                              cut_off = CUT_OFF[0]):
    
    
    ind_names = get_individualnames()

    if title_string in ind_names.keys():
        title_string = ind_names[title_string]
        
    ensure_directory(output_fig)
    
    wb = load_workbook(xlsx_file)
    ws = wb['Verhalten']
    
    ws = read_worksheet(wb, 'Verhalten')
    alphabet = get_alphabet()
    
        
    max_row = max(ws.keys())
    maximum_row = min(max_row, cut_off+1)
    
    max_col = np.max( [ len( ws[i] ) for i in ws.keys() ] ) 
    nights = max_col - 3
    
    ws_avg = wb['Overview']
    ws_avg = read_worksheet(wb, 'Overview')
    
    avgs = {}
    sems = {}
    for i in draw.keys():
        avgs[i] = float(ws_avg[3][alphabet[11+i]])                     
        sems[i] = float(ws_avg[4][alphabet[11+i]])           
    percentages = {}     
    for i in draw.keys():
        percentages[i] = [] 
    
    for zeile in range(2, maximum_row):
        tmp = [ int( ws[zeile][alphabet[i]] ) for i in range(3, max_col) ]
        c = Counter(tmp)
        for i in draw.keys():          
            percentages[i].append(float(  c[i] / nights  ))
            
    for i in draw.keys():
        percentages[i] = running_mean(percentages[i], 9)
    
    plt.figure(figsize=(24,6))
    curves = []
    legend_elements = []
    for h in draw.keys(): 
        draw_vals = [ percentages[h][j] for j in range(len(percentages[h])) ]
        draw_upper = [ min(percentages[h][j] + percentages[h][j]*(1-percentages[h][j])/np.sqrt(nights), 0.999) for j in range(len(percentages[h])) ]
        draw_lower = [ max(percentages[h][j] - percentages[h][j]*(1-percentages[h][j])/np.sqrt(nights), 0.001) for j in range(len(percentages[h])) ]
        x, = plt.plot(draw_vals, color=draw[h], linestyle = '-', 
                 linewidth=0.75, label = str(h))
        curves.append(x)
        legend_elements.append(x)
        
        u, = plt.plot(draw_upper, color=draw[h], linestyle = ':', alpha = 0.2,
                     linewidth=0.3, label = 'u'+str(h))
        curves.append(u)
        
        l, = plt.plot(draw_lower, color=draw[h], linestyle = ':', alpha = 0.2,
                     linewidth=0.3, label = 'l'+str(h))
        curves.append(l)
        
        
        x_coord = np.array([j for j in range(len(draw_upper))], dtype='float32')        
        plt.fill_between(x_coord, draw_lower, draw_upper, facecolor = draw[h], alpha = 0.2, interpolate = True)
        
        draw_avg = [avgs[h] for j in range(maximum_row-2)]
        sem_plus = [avgs[h] + sems[h] for j in range(maximum_row-2)]
        sem_minus = [avgs[h] - sems[h] for j in range(maximum_row-2)]
        
        plt.plot(draw_avg, color=draw[h], linestyle='-', linewidth=1)
        plt.plot(sem_plus, color=draw[h], linestyle=':', linewidth=1)
        plt.plot(sem_minus, color=draw[h], linestyle=':', linewidth=1)
    
    plt.ylabel('Percentage', fontsize=12)
    plt.xlabel('Time', fontsize=12)
    plt.xticks(x_ticks, x_labels, fontsize=12)
    plt.yticks([np.round(x/10,1) for x in range(0, 11) ], fontsize=12)
    
    plt.legend( legend_elements, [behavs[h] for h in draw.keys()], loc='upper left', bbox_to_anchor=(0.78, 1.1),
          ncol=3, fancybox=True, shadow=True, fontsize=11 )
    plt.title(title_string, fontsize=14)
    
    axes = plt.gca()
    axes.margins(x=0.005)
    axes.set_ylim([-0.02, 1.02])
    
    plt.savefig(output_fig)
    plt.close()


def _merge_predictions(input_folder, output_xlsx, behavs, counter_j, output_fig, draw_behavs):
    
    def write_standard_sheet(dictionary, wb, title, behavs = behavs):
        
        ws = wb.create_sheet(title)
        
        for i in range(len(behavs)):
            ws.cell(1,i+2).value = behavs[i]
        
        zeile = 2
        for datum in sorted(dictionary.keys()):
            ws.cell(zeile, 1).value = datum
            for i in range(len(behavs)):
                ws.cell(zeile, i+2).value = dictionary[datum][i]
            zeile += 1
    
    def _write_sequences(wb, seq_dict, interval_labels):
        
        ws = wb.create_sheet('Verhalten')
        ws.cell(1,1).value = 'Intervall'
        ws.cell(1,2).value = 'Startzeit'
        ws.cell(1,3).value = 'Endzeit'
        
        dates = sorted(seq_dict.keys())

        for i in range(len(dates)):
            ws.cell(1, i+4).value = dates[i]
        
        for j in range(len(interval_labels[0])):            
            for h in range(3):
                ws.cell(j+2, h+1).value = interval_labels[h][j]
            
        for i in range(len(dates)):
            zeile = 2
            for j in range(len(seq_dict[dates[i]])):
                ws.cell(zeile, i+4).value = seq_dict[dates[i]][j]
                zeile += 1                      
        
    
    
    def _write_overview(ws, amounts, durations, percentages, phases, behavs = behavs):
        
        
        def _write_stats(ws, values_in, spalte, behavs = behavs, prec = 3):
            
            values = {}
            
            for j in range(len(behavs)):
                values[j] = []
            
            for h in range(len(behavs)):
                for val in values_in.values():                
                    values[h].append(val[h])
                    
            for h in range(len(behavs)):
                
                total = []
                if type(values[h][0]) == list:
                    for i in values[h]:
                        total += i
                else:
                    total = values[h]
                    
                ws.cell(2, spalte + h).value = behavs[h]
                ws.cell(3, spalte + h).value = np.round(np.mean(total), prec)
                ws.cell(4, spalte + h).value = np.round(SEM(total), prec)
                
                ws.cell(6, spalte + h).value = np.round(np.median(total), prec)
                ws.cell(7, spalte + h).value = np.round(np.quantile(total, 0.25), prec)
                ws.cell(8, spalte + h).value = np.round(np.quantile(total, 0.75), prec)
        
        
        ws.cell(3,1).value = 'Mean'
        ws.cell(4,1).value = 'SEM'       
        
        ws.cell(6,1).value = 'Median'
        ws.cell(7,1).value = '0.25-quantile'
        ws.cell(8,1).value = '0.75-quantile'
        
        
        spalte = 2
        ws.cell(1, spalte).value = 'Anzahl Phasen'
        _write_stats(ws, amounts, spalte, prec=2)
        
        spalte += len(behavs) + 1
        ws.cell(1, spalte).value = 'Dauer (gesamt [sec])'
        _write_stats(ws, durations, spalte, prec=1)
        
        spalte += len(behavs) + 1
        ws.cell(1, spalte).value = 'Anteil'
        _write_stats(ws, percentages, spalte, prec=4)
        
        spalte += len(behavs) + 1
        ws.cell(1, spalte).value = 'Durchschnittliche Dauer [sec]'
        _write_stats(ws, phases, spalte, prec=1)
        
 
            
        
    
    def _write_phases(wb, phase_dict, behavs = behavs):
        ws = wb.create_sheet('Phasen')

        for h in range(len(behavs)):
            ws.cell(1, h+1).value = behavs[h]
        
        
        phasen_joint = {}
        for j in range(len(behavs)):
            phasen_joint[j] = []
        
        for datum, phasenliste in phase_dict.items():
            for h in range(len(behavs)):
                phasen_joint[h] += phasenliste[h]
        
        for h in range(len(behavs)):
            zeile = 2
            for phase in phasen_joint[h]:
                ws.cell(zeile, h+1).value = phase
                zeile += 1
                
        
    files = sorted([ input_folder + f for f in os.listdir(input_folder) if f.endswith(".xlsx") ])
    
    phases = {}
    avg_phases = {}
    sequences = {}
    amounts = {}
    durations = {}
    percentages = {}
    print("***************************")
    for xls in files:
        
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        print(current_time)
        
        print("Currently processing " + xls)
        filename_xls = xls.split("/")[-1]
        datum = filename_xls.split("_")[0]
        anzahl, dauer, sequence, phase_lists = _xlsx_to_prediction(xls)
        
        phases[datum] = phase_lists
        sequences[datum] = sequence
        amounts[datum] = anzahl
        durations[datum] = dauer
        
        avg_phases[datum] = []
        
        for j in range(len(behavs)):
            avg_phases[datum].append( np.round(np.mean(phase_lists[j]), 1) )
        
        total_duration = np.sum(dauer)
        percentages[datum] = [np.round(dauer[i] / total_duration,4) for i in range(len(dauer))]
    
    print("***************************")
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    print(current_time)
    print("Creating merged file.")
        
    wb = Workbook()
    ws = wb.active 
    ws.title = 'Overview'
    
    
    
    
    _write_overview(ws, amounts, durations, percentages, avg_phases)
    print("- Successfully created Overview-sheet. ")
    
    write_standard_sheet(dictionary = amounts, wb = wb, title = 'Anzahl')
    print("- Successfully created Anzahl-sheet. ")
    write_standard_sheet(dictionary = durations, wb = wb, title = 'Dauer')
    print("- Successfully created Dauer-sheet. ")
    write_standard_sheet(dictionary = percentages, wb = wb, title = 'Anteil')
    print("- Successfully created Anteil.")
    
    write_standard_sheet(dictionary = avg_phases, wb = wb, title = 'Durchschnittliche Phasen')
    print("- Successfully created Durchschnittliche Phasen.")
    
    interval_labels = _get_interval_labels(files[0])
    _write_sequences(wb, sequences, interval_labels)
    print("- Successfully created interval overview-sheet.")
    
    _write_phases(wb, phases)
    print("- Successfully accumulated phases.")
    
    ensure_directory(output_xlsx)
    wb.save(output_xlsx)
   
    create_timeline_from_xlsx(xlsx_file = output_xlsx, 
                              output_fig = output_fig, 
                              draw = draw_behavs,
                              x_labels = TIMES[counter_j], 
                              x_ticks = CORRESPONDING_INTERVALS[counter_j],
                              behavs = behavs,
                              title_string = GRAPHIC_TITLE[counter_j],
                              cut_off = CUT_OFF[counter_j])


def draw_timelines( basefolder = OUTPUT_FOLDER_XLSX, individualcodes = INDIVIDUALS_TO_MERGE, ext = EXT ):
    input_xlsx = []
    for individual in individualcodes:
        art, zoo, ind = individual.split('_')
        xlsx = basefolder + art + '/' + individual + '_Übersicht_Stat'+ext+'.xlsx'
        input_xlsx.append(xlsx)
    
    for counter_j in range(len(input_xlsx)):
        create_timeline_from_xlsx(xlsx_file = input_xlsx[counter_j], 
                              output_fig = TIMELINE[counter_j], 
                              x_labels = TIMES[counter_j], 
                              x_ticks = CORRESPONDING_INTERVALS[counter_j],
                              title_string = GRAPHIC_TITLE[counter_j],
                              cut_off = CUT_OFF[counter_j])

def merge_predictions(input_folders = CONTAINING_FOLDER, 
                      output_xlsxs = OUTPUT_XLSX, 
                      behavs = BEHAVIORS, 
                      output_figs = TIMELINE,
                      draw_behavs = DRAW_BEHAVS):

    for j in range(NUMBER_INDIVIDUALS):
        print("******************************")
        print('Current Individual: {}, current zoo: {}, current species: {}'.format(INDIVIDUAL[j], ZOO[j], SPECIES[j]))
        _merge_predictions(input_folder = input_folders[j], 
                           output_xlsx = output_xlsxs[j], 
                           behavs = behavs, 
                           counter_j = j, 
                           output_fig = output_figs[j], 
                           draw_behavs = draw_behavs)
