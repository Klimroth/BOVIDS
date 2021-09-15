#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "Max Hahn-Klimroth"
__copyright__ = "Copyright 2020, M. Hahn-Klimroth, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes"]
__license__ = "MIT"
__version__ = "1.0"
__maintainer__ = "M. Hahn-Klimroth"
__status__ = "Development"

import configuration as cf
import csv, os, copy
import numpy as np
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
from matplotlib.collections import PolyCollection



def _extract_single_phases(np_array, timeinterval = cf.INTERVAL_LENGTH):
    """
    

    Parameters
    ----------
    np_array : TYPE
        DESCRIPTION.
    timeinterval : TYPE, optional
        DESCRIPTION. The default is cf.INTERVAL_LENGTH.

    Returns
    -------
    phases : [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ]

    """
    phases = []
    LastBehav = np_array[0]
    iCurrLen = 1
    j = 1
    start_interval = 1

    for DistLine in np_array:
        
        if LastBehav == DistLine:
            iCurrLen += 1
        else:
            phases.append([iCurrLen*timeinterval, LastBehav, start_interval, j-1])
            iCurrLen = 1
            start_interval = j
            LastBehav = DistLine
        j += 1

    phases.append([iCurrLen*timeinterval, LastBehav, phases[-1][3] + 1, phases[-1][3] + 1 + iCurrLen])
    return phases
    
    

def postprocess_night(input_sequence, 
                       behavior_mapping,
                       post_processing_rules,
                       current_video_start,
                       current_video_end,
                       behavior_names = cf.BEHAVIOR_NAMES,
                       apply_postprocessing = True,
                       add_observation_time = True):
    
    """
    Inputfile: Either a .xlsx in the AI style or alternatively a csv produced by the BORIS prediction.
    In both cases it is in 7s intervals.
    
    behavior_mapping: mapping which should be applied (for instance, map sleep to lying)
    
    post_processing_rules: set of rules to be applied, rolling average will be dismissed.
    
    """
    
    
       
    
    
    
                
    
    
    def _join_phases(phase_list):
        """        
            Joins consecutive phases of the same behavior to a single phase.

        """
    
        ret = [phase_list[0]]
        
        for j in range(1, len(phase_list)):
            if phase_list[j][1] == phase_list[j-1][1]:
                ret[-1][0] += phase_list[j][0]
                ret[-1][3] = phase_list[j][3]
            else:
                ret.append(phase_list[j])
        
        return ret
    
    
    def _remove_short_phases(phase_list, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Removes those short phases which are very unlikely due to configuration
        """
        
        def _check_first_phase(phase_list):
            
            if len(phase_list) <= 1:
                return False
            
            l = phase_list[0][0]
            if l < 9*interval_len:
                phase_list[0][1] = phase_list[1][1]
                return True
            
            return False
            
        phase_list = _join_phases(phase_list)
        
        if len(phase_list) < 3:
            return phase_list
        
        
        changes_done = True
        
        while changes_done:
            
            last_behavior = phase_list[0][1]
            current_behavior = None
            next_behavior = None
            
            changes_done = False
            changes_done = _check_first_phase(phase_list)
            
            for j in range(1, len(phase_list)-1):
                last_behavior = phase_list[j-1][1]
                current_behavior = phase_list[j][1]
                next_behavior = phase_list[j+1][1]
                
                # lying
                if [last_behavior, current_behavior, next_behavior] == [2,1,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SLS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,0]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_ALA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_ALS']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,1,0]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SLA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,1,0]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OLA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_ALO']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,1,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OLS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,1,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SLO']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                        
                        
                # standing
                elif [last_behavior, current_behavior, next_behavior] == [2,0,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SAS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LAL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_LAS']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,0,1]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SAL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,0,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SAO']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,0,1]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OAL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_LAO']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,0,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OAS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                        
                        
                        
                # sleeping
                elif [last_behavior, current_behavior, next_behavior] == [0,2,0]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_ASA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,2,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_ASL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LSL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,0]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LSA']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True                
                elif [last_behavior, current_behavior, next_behavior] == [0,2,3]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_ASO']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,2,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_OSL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,2,0]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_OSA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,3]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LSO']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                        
                # truncation
                elif current_behavior == 4: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_TRUNCATION']*interval_len:
                        phase_list[j][1] = last_behavior
                        changes_done = True         
        
            phase_list = _join_phases(phase_list)
            
        return phase_list

    def _remove_out(phase_list, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """
        Removes those phases of behavior out which are shorter than the value given by the configuration
        """
        j = 0
        for phase in phase_list:
            if phase[1] == 3 and j >= 1: # it is out
                if phase[0] <= pp_rules['MIN_LEN_OUT']*interval_len and phase[2] > pp_rules['MIN_TIME_OUT']:
                    phase[1] = phase_list[j-1][1]
            j += 1
        
        return phase_list
    
     
    
    def _get_time(interval_num, video_start = cf.GLOBAL_STARTING_TIME, interval_len = cf.INTERVAL_LENGTH):
        """
        
    
        Parameters
        ----------
        interval_num : TYPE
            DESCRIPTION.
        video_start : TYPE, optional
            DESCRIPTION. The default is GLOBAL_STARTING_TIME (when the standard observation should start).
    
        Returns
        -------
        startzeit, endzeit, startframe, endframe
    
        """
        startframe = (interval_num-1)*interval_len + 1
        endframe = interval_num*interval_len
        
        start_hours = startframe // (60*60)
        start_minutes = (startframe % (60*60)) // 60
        start_seconds = startframe % 60
        
        end_hours = endframe // (60*60)
        end_minutes = (endframe % (60*60)) // 60
        end_seconds = endframe % 60
        
        start_time = str((video_start + start_hours)%24).zfill(2) + ":" + str(start_minutes).zfill(2) + ":" + str(start_seconds).zfill(2)
        end_time = str((video_start + end_hours)%24).zfill(2) + ":" + str(end_minutes).zfill(2) + ":" + str(end_seconds).zfill(2)
        
        return start_time, end_time, startframe, endframe
    
    def _get_sparse_intervals_from_phases(phases):
       
        new_intervallist = [] # sequence of behaviors
        for j in range(1, len(phases)+1):
            curr_phase = phases[j-1]
            curr_start = curr_phase[2]
            curr_end = curr_phase[3]
            curr_dur = curr_end - curr_start
            for i in range(curr_dur + 1):
                new_intervallist.append(curr_phase[1])
        return new_intervallist
     
    
 
    
    def _remove_out_fluctuation(phase_list, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Removes those fluctuations between out and lying or standing if out is not too long and the rest is sufficiently long
        """
        
        def _get_current_run_endindex(start_index, phase_list, max_out):
            actual_behavior = phase_list[start_index-1][1]
            is_run = True
            j = start_index
            total_dur_actual_behavior = phase_list[start_index-1][0]
            total_dur_out = 0
            while is_run and j < len(phase_list):
                curr_behav = phase_list[j][1]
                if curr_behav == actual_behavior:
                    total_dur_actual_behavior += phase_list[j][0]
                    j += 1
                elif curr_behav == 3 and phase_list[j][0] <= max_out:
                    total_dur_out += phase_list[j][0]
                    j += 1
                else:
                    is_run = False
            
            return j, total_dur_actual_behavior, total_dur_out
                    
                    
            
        max_out = post_processing_rules['OUT_FLUCTUATION_REMOVAL_MAX']*interval_len
        min_actual_behavior = post_processing_rules['OUT_FLUCTUATION_REMOVAL_MIN_BEHAV']
        

        j = 1
        changes_done = False
        while j < len(phase_list)-1:
            last_behavior = phase_list[j-1][1]
            current_behavior = phase_list[j][1]
            next_behavior = phase_list[j+1][1]

            if current_behavior == 3:
                if phase_list[j][0] > max_out:
                    j += 1
                    continue
                if next_behavior != last_behavior:
                    j += 1
                    continue
                end_index, dur_run, dur_out = _get_current_run_endindex(j, phase_list, max_out)
                if dur_run*min_actual_behavior <= dur_out:
                    j += 1
                    continue
                
                changes_done = True
                for h in range(j, end_index):
                    phase_list[h][1] = last_behavior
                j = end_index + 1
            else:
                j += 1
        
        phase_list = _join_phases(phase_list)
        

        return phase_list, changes_done
                    
                
 
    
    def _get_video_length(start, end):
            if end > start:
                return end-start
            
            return 24 + (end-start)
        
    def _add_observation_time(multiple_frame_dist, vid_start = current_video_start, 
                          vid_end = current_video_end,
                          observation_start = cf.GLOBAL_STARTING_TIME,
                          observation_end = cf.GLOBAL_ENDING_TIME,
                          interval = cf.INTERVAL_LENGTH):
        
        
        observation_intervals = _get_video_length(observation_start, observation_end)*3600 // interval 
        intervals_per_hour = 3600//interval
        
        ret = []
        frames_fill_start = (vid_start-observation_start) * intervals_per_hour
        frames_fill_end = (observation_end - vid_end) * intervals_per_hour
        

        if frames_fill_start > 0:
            for i in range(frames_fill_start):
                ret.append( 3 )
        
        for j in range(len(multiple_frame_dist)):
            ret.append( multiple_frame_dist[j] )
        
        if frames_fill_end > 0:
            for i in range(frames_fill_end+1):
                ret.append( 3 )
        
        missing_intervals = observation_intervals - len(ret)
        if missing_intervals > 0:
            for j in range(missing_intervals):
                ret.append( 3 )
        return ret

    ensemble_sparse = copy.deepcopy(input_sequence)
    if add_observation_time:
        ensemble_sparse =_add_observation_time(ensemble_sparse)
        
    if apply_postprocessing:
        ensemble_sparse = _extract_single_phases( ensemble_sparse )
        #print('orig')
        #print(ensemble_sparse)
        ensemble_sparse = _remove_out(ensemble_sparse)
        #print('out removal')
        #print(ensemble_sparse)
        ensemble_sparse = _remove_short_phases(ensemble_sparse)
        #print('short')
        #print(ensemble_sparse)        
        x = True
        while x:            
            ensemble_sparse, x = _remove_out_fluctuation(ensemble_sparse)
            #print('fluc loop')
            #print(ensemble_sparse)
        ensemble_sparse = _remove_short_phases(ensemble_sparse)
        sparse_postprocessed = _get_sparse_intervals_from_phases(ensemble_sparse)
        
        return sparse_postprocessed
    
    return ensemble_sparse
    

def write_output_xlsx_file(ensemble_sparse, outputfolder, datum, individual_code, mode):
    
    def _get_sparse_intervals_from_phases(phases):
       
        new_intervallist = [] # sequence of behaviors
        for j in range(1, len(phases)+1):
            curr_phase = phases[j-1]
            curr_start = curr_phase[2]
            curr_end = curr_phase[3]
            curr_dur = curr_end - curr_start
            for i in range(curr_dur + 1):
                new_intervallist.append(curr_phase[1])
        return new_intervallist
     
    
    def _write_xlsx_statistics(phases_ordered, phases_unordered, interval_list, outputfolder, 
                               filename, extension, names_behav = cf.BEHAVIOR_NAMES):
        wb = Workbook()
        ws = wb.active
        ws.title = "Zeitintervalle"
        
        ws["A1"] = "Zeitintervall"
        ws["B1"] = "Startzeit"
        ws["C1"] = "Endzeit"
        ws["D1"] = "Startframe"
        ws["E1"] = "Endframe"
        ws["F1"] = "Verhaltenscode"
        ws["G1"] = "Verhaltenswort"
        
        
        for j in range(len(interval_list)):
            starttime, endtime, startframe, endframe = _get_time(j+1)
            ws["A"+str(j+2)] = j+1
            ws["B"+str(j+2)] = starttime
            ws["C"+str(j+2)] = endtime
            ws["D"+str(j+2)] = startframe
            ws["E"+str(j+2)] = endframe
            ws["F"+str(j+2)] = interval_list[j]
            ws["G"+str(j+2)] = names_behav[interval_list[j]]
            
        
                    
        #  [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ]
        # phases: [ [dauer (sek), verhaltenscode, startinterval, endinterval] ] (stat sheet 2)
        ws2 = wb.create_sheet("Aktivitätsphasen_geordnet", -1)
        
        ws2["A1"] = "Phase"
        ws2["B1"] = "Nummer der Phase des Verhaltens"
        ws2["C1"] = "Startzeit"
        ws2["D1"] = "Endzeit"
        ws2["E1"] = "Startintervall"
        ws2["F1"] = "Endintervall"
        ws2["G1"] = "Länge [sec]"
        ws2["H1"] = "Verhaltenscode"
        ws2["I1"] = "Verhaltenswort"
        
        j = 0
        amount_phases = [0]*len(names_behav)
        for phase in phases_unordered:
            amount_phases[phase[1]] += 1
            starttime, _, _, _ = _get_time(phase[2])
            _, endtime, _, _ = _get_time(phase[3])
            
            ws2["A"+str(j+2)] = j+1
            ws2["B"+str(j+2)] = amount_phases[phase[1]] 
            ws2["C"+str(j+2)] = starttime
            ws2["D"+str(j+2)] = endtime
            ws2["E"+str(j+2)] = phase[2]
            ws2["F"+str(j+2)] = phase[3]
            ws2["G"+str(j+2)] = phase[0]
            ws2["H"+str(j+2)] = phase[1]
            ws2["I"+str(j+2)] = names_behav[phase[1]]
            
            j += 1
            
            
        # phases_ordered [ [ [dauer (s), 0, startinterval, endinterval] ... ], [dauer (s), 1, startinterval, endinterval], ... ] (stat sheet 3) 
        ws3 = wb.create_sheet("Aktivitätsphasen_Übersicht", -1)
        colname=["A", "B", "C", "D", "E", "F"]
        
        for i in range(4):
            ws3[colname[i]+"1"] = names_behav[i]
            j = 0
            for phase_stand in phases_ordered[i]:
                ws3[colname[i]+str(j+2)] = phase_stand
                j += 1
        
        
        
        
        ws4 = wb.create_sheet("Statistik", -1)
        
        ws4["A2"] = "Anzahl Phasen"
        ws4["A3"] = "Gesamtdauer [sec]"
        ws4["A4"] = "Anteil [%]"
        
        ws4["A6"] = "Median [sec]"
        ws4["A7"] = "0.25-Quantil"
        ws4["A8"] = "0.75-Quantil"
        ws4["A10"] = "Mean [sec]"
        ws4["A11"] = "SEM"
        
        total_duration = 0
        for j in range(len(names_behav)):
            total_duration += np.sum(phases_ordered[j])
        for j in range(len(names_behav)):
            amount = len(phases_ordered[j])
            ws4[colname[j+1]+"1"] = names_behav[j]
            if amount > 0:
                ws4[colname[j+1]+"2"] = amount
                ws4[colname[j+1]+"3"] = np.sum(phases_ordered[j])
                ws4[colname[j+1]+"4"] = round(100 * np.sum(phases_ordered[j]) / total_duration, 1)
                
                ws4[colname[j+1]+"6"] = np.median(phases_ordered[j])
                ws4[colname[j+1]+"7"] = np.quantile(phases_ordered[j], 0.25)
                ws4[colname[j+1]+"8"] = np.quantile(phases_ordered[j], 0.75)
                ws4[colname[j+1]+"10"] = np.mean(phases_ordered[j])
                ws4[colname[j+1]+"11"] = np.std(phases_ordered[j]) / np.sqrt(amount)
            else:
                ws4[colname[j+1]+"2"] = amount
                ws4[colname[j+1]+"3"] = 0
                ws4[colname[j+1]+"4"] = 0
                
                ws4[colname[j+1]+"6"] = 0
                ws4[colname[j+1]+"7"] = 0
                ws4[colname[j+1]+"8"] = 0
                ws4[colname[j+1]+"10"] = 0
                ws4[colname[j+1]+"11"] = 0
        
        ensure_dir(outputfolder)
        wb.save(outputfolder + filename + extension + '.xlsx')
    
    def ensure_dir(directory):
        if not os.path.exists(directory):
            os.makedirs(directory)
    
    def _get_time(interval_num, video_start = cf.GLOBAL_STARTING_TIME, interval_len = cf.INTERVAL_LENGTH):
        """
        
    
        Parameters
        ----------
        interval_num : TYPE
            DESCRIPTION.
        video_start : TYPE, optional
            DESCRIPTION. The default is GLOBAL_STARTING_TIME (when the standard observation should start).
    
        Returns
        -------
        startzeit, endzeit, startframe, endframe
    
        """
        startframe = (interval_num-1)*interval_len + 1
        endframe = interval_num*interval_len
        
        start_hours = startframe // (60*60)
        start_minutes = (startframe % (60*60)) // 60
        start_seconds = startframe % 60
        
        end_hours = endframe // (60*60)
        end_minutes = (endframe % (60*60)) // 60
        end_seconds = endframe % 60
        
        start_time = str((video_start + start_hours)%24).zfill(2) + ":" + str(start_minutes).zfill(2) + ":" + str(start_seconds).zfill(2)
        end_time = str((video_start + end_hours)%24).zfill(2) + ":" + str(end_minutes).zfill(2) + ":" + str(end_seconds).zfill(2)
        
        return start_time, end_time, startframe, endframe
    
    def _draw_timeline(phases_unordered, save_path, title, extension, behav_names2 = cf.BEHAVIOR_NAMES, 
                       colormapping = cf.COLOR_MAPPING):
        
        behav_names = copy.deepcopy(behav_names2)
        data = []
        for phase in phases_unordered:
            startframe = phase[2]
            endframe = phase[3]
            behavior = behav_names[phase[1]]
            
            data.append((startframe, endframe, behavior))
        
        cats ={}
        for j in range(len(behav_names)):
            cats[behav_names[j]] = (j+1)/2
        
        verts = []
        colors = []
        for d in data:
            v =  [(d[0], cats[d[2]]-.075),
                  (d[0], cats[d[2]]+.075),
                  (d[1], cats[d[2]]+.075),
                  (d[1], cats[d[2]]-.075),
                  (d[0], cats[d[2]]-.075)]
            verts.append(v)
            colors.append(colormapping[d[2]])
        
        bars = PolyCollection(verts, facecolors=colors)
        
        x_ticks_set = []
        x_ticks_labels = []
        for j in range( int(cf.GLOBAL_VIDEO_LENGTH / cf.INTERVAL_LENGTH) - 1):
            if j % int(3600/cf.INTERVAL_LENGTH) == 0:
                start_time, end_time, startframe, endframe = _get_time(j)
                hh, mm, ss = start_time.split(":")
                hh = str(int(hh)+1).zfill(2)
                mm = ':00'
                
                x_ticks_labels.append(hh + mm)
                x_ticks_set.append(j)
            
        
        
        
        fig = plt.figure(figsize=(13,4))
        ax = fig.add_subplot(111)
        
        fig.autofmt_xdate(rotation=60)
        fig.suptitle(title, fontsize=16)
        
        ax.add_collection(bars)
        ax.autoscale()
        
        ax.set_xticks( x_ticks_set )
        ax.set_xticklabels( x_ticks_labels )
        ax.set_yticks([0.5, 1, 1.5, 2, 2.5, 3])
        behav_names.append(" ")
        ax.set_yticklabels(behav_names)
        
      
        plt.savefig(save_path + extension + '.png')
        plt.close()
        
    
    ensemble_sparse = _extract_single_phases( ensemble_sparse )
    phases_ordered = [ [x[0] for x in ensemble_sparse if x[1] == 0], 
                          [x[0] for x in ensemble_sparse if x[1] == 1],
                          [x[0] for x in ensemble_sparse if x[1] == 2],
                          [x[0] for x in ensemble_sparse if x[1] == 3],
                          [x[0] for x in ensemble_sparse if x[1] == 4]
                         ]
    sparse_postprocessed = _get_sparse_intervals_from_phases(ensemble_sparse)
        
    
    if mode == 'binary':
        ext = '_binary'
    else:
        ext = ''
    filename = datum + '_' + individual_code + '_statistics' 
    _write_xlsx_statistics(phases_ordered, ensemble_sparse, sparse_postprocessed, outputfolder, filename, ext)
        
    save_path = outputfolder + datum + '_' + individual_code + '_timeline'
    title = datum + '_' + individual_code
    _draw_timeline(ensemble_sparse, save_path, title, ext)