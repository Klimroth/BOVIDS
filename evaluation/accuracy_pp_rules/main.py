# !/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = "Max Hahn-Klimroth"
__copyright__ = "Copyright 2021, M. Hahn-Klimroth, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes", "T. Kapetanopoulos"]
__license__ = "MIT"
__version__ = "1.0"
__maintainer__ = "M. Hahn-Klimroth"
__status__ = "Development"

"""

"""



from configuration import BEHAVIOR_NAMES as BEHAVIORS
from configuration import ERROR_TOLERANCE_INTERVALS, get_pp_rule, get_vid_start_end
from configuration import INTERVAL_LENGTH, CONF_SPALTEN, CONF_ZEILEN

from apply_postprocessing import postprocess_night, write_output_xlsx_file
from get_accuracy import compare_one_night, describe_errors

import numpy as np
from openpyxl import Workbook, load_workbook
from sklearn.metrics import precision_recall_fscore_support, accuracy_score, confusion_matrix
from scipy.stats import sem as StandardError
import csv, os, string
import pandas as pd



"""
Generate BOVIDS xlsx files out of the real data using the post-processing rules already determined.
generate_bovids_files() 
"""
INDIVIDUALCODES_BOVIDS = []
CSV_FOLDER = ''
OUTPUT_FOLDER_BOVIDS = '' 

"""
Generate accuracy sheets.
generate_accuracy()
"""

USE_SPECIES = True # if True, SPEC_ACCURACY is used, otherwise, INDIVIDUALCODES_ACCURACY.

SPEC_ACCURACY = []
INDIVIDUALCODES_ACCURACY = []


BASE_FOLDER_CSV = 
BASE_FOLDER_AI_PREDICTION =
OUTPUT_FOLDER_ACCURACY = 



COMPARE_REAL_WITH_AI = True
COMPARE_REAL_WITH_REALPP = True
COMPARE_REALPP_WITH_AI = True

TOTAL = True
BINARY = True


"""
Generate overview file from contained accuracy sheets
necessary structure of the folder: .../species_zoo_individual/different_accuracy_sheets.xlsx
creates for each type of comparision a sheet, consisting of two tabs (total, binary)
generate_overview() 
"""

INPUT_FOLDER_FOR_OVERVIEW = ''
OUTPUT_FOLDER_OVERVIEW = ''



#############################################################################
# general code
#############################################################################


def get_zoospec_list(speclist = SPEC_ACCURACY, auswertungsordner = BASE_FOLDER_CSV):
    ret = []
    for spec in speclist:
        if not os.path.exists(auswertungsordner + spec):
            continue
        for zoo in os.listdir(auswertungsordner + spec):
            if not os.path.isdir( auswertungsordner + spec + '/' + zoo ):
                continue
            ret.append( spec + '_' + zoo )
    return ret

ZOOSPEC_ACCURACY = get_zoospec_list()


def get_individualcodelist(zoospec = ZOOSPEC_ACCURACY, auswertungsordner = BASE_FOLDER_CSV):
    
    def get_individualcodes(f):
        inds = []
        files = [fn for fn in sorted(os.listdir(f)) if fn.endswith('SUM-7s_pred.csv')]
        
        for fn in files:
            date, spec, zoo, ind, _, _ = fn.split('_')
            indcode = spec + '_' + zoo + '_' + ind
            if not indcode in inds:
                inds.append(indcode)
        return inds
    
    indlist = []
    for speczoo in zoospec:
        spec, zoo = speczoo.split('_')
        csv_folder = '{}{}/{}/Auswertung/Boris_KI/csv-Dateien/'.format(auswertungsordner, spec, zoo)
        individualcodes = get_individualcodes(csv_folder)
        for individualcode in individualcodes:
            indlist.append(individualcode)
    return indlist

if USE_SPECIES:
    INDIVIDUALCODES_ACCURACY = get_individualcodelist()


def _get_speczoo(l):
    spec, zoo, ind = [], [], []
    for indcode in l:
        s, z, i = indcode.split('_')
        spec.append(s)
        zoo.append(z)
        ind.append(i)
    return spec, zoo, ind


SPECIES, ZOO, INDIVIDUAL = _get_speczoo(INDIVIDUALCODES_ACCURACY)


def ensure_directory(filename):
    path = os.path.dirname(os.path.abspath(filename))
    if not os.path.exists(path):
        os.makedirs(path)

#############################################################################
# Code for generating the accuracy sheets #
#############################################################################

def get_input_folders(mode, j,
        csv_base = BASE_FOLDER_CSV,
        ai_base = BASE_FOLDER_AI_PREDICTION,
        species = SPECIES,
        zoo = ZOO,
        ind = INDIVIDUAL
        ):
    
    if mode == 'binary' or mode == 'total':
        csv_folder = '{}{}/{}/Auswertung/Boris_KI/csv-Dateien/'.format(csv_base, species[j], zoo[j])
        xlsx_folder = '{}{}/{}/{}/{}/final/'.format(ai_base, species[j], zoo[j], ind[j], mode)
        
        return csv_folder, xlsx_folder
    
    return [], []




def get_behavior_mapping_and_extensions(mode):

    if mode == 'total':
        return {0:0, 1:1, 2:2, 3:3, 4:4}, ''
    elif mode == 'binary':
        return {0:0, 1:1, 2:1, 3:3, 4:4}, '_binary'

def get_output_xlsx(mode, compare_mode, j, species = SPECIES, zoo = ZOO, ind = INDIVIDUAL, out = OUTPUT_FOLDER_ACCURACY):
    
    return '{}{}_{}_{}/{}_{}_{}_Accuracy_{}-{}.xlsx'.format(out, species[j], zoo[j], ind[j], species[j], zoo[j], ind[j], compare_mode, mode)





def get_dates_to_evaluate(csv_folder, j, spec = SPECIES, zoo = ZOO, ind = INDIVIDUAL):
    
    def _get_video_dates(individual_code, csv_folder):
        csv_files = [f for f in sorted(os.listdir(csv_folder)) if f.endswith('.csv')]
        ret = []
        for f in csv_files:
            date, species, zoo, individual, _, _ = f.split('_')
            if '{}_{}_{}'.format(species, zoo, individual) == individual_code:
                ret.append(date)
        return ret


    ind_code = '{}_{}_{}'.format(spec[j], zoo[j], ind[j])
    csv_f = csv_folder
    dates = _get_video_dates(ind_code, csv_f)
    
    return dates

       

        

def write_outputfile(worksheet_overview, 
                     list_of_errors, 
                     conf_mat, 
                     out, 
                     behaviors = BEHAVIORS, 
                     sensitivity = ERROR_TOLERANCE_INTERVALS,
                     label_real = 'real',
                     label_ai = 'AI'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overview'
    
    ensure_directory(out)
    
    kennzahlen = ['Anzahl_Phasen_', 'Anteil_', 'Median_Length_']
    # Überschriften
    spalte = 2
    for kennzahl in kennzahlen:
        for behavior in behaviors:
            ws.cell(1, spalte).value = kennzahl+behavior
            ws.cell(2, spalte).value = label_real
            ws.cell(2, spalte + 1).value = label_ai       
            spalte += 2
        spalte += 1
        
    for behavior in behaviors:
        ws.cell(1, spalte).value = 'F-Score'
        ws.cell(2, spalte).value = behavior  
        spalte += 1
    
    spalte += 1
    
    ws.cell(1, spalte).value = 'accuracy'
 
        
    zeile = 3
    
    for datum in sorted(worksheet_overview.keys()):
        
        if datum == 'AVG':
            zeile += 1
        
        ws.cell(zeile, 1).value = datum
        
        spalte = 2
        for kennzahl in kennzahlen:
            for behavior in behaviors:
                    
                ws.cell(zeile, spalte).value = worksheet_overview[datum][kennzahl+behavior][0]
                ws.cell(zeile, spalte + 1).value = worksheet_overview[datum][kennzahl+behavior][1]      
                spalte += 2
            spalte += 1
        
        for behavior in behaviors:
            ws.cell(zeile, spalte).value = worksheet_overview[datum]['fscore_'+behavior]            
            spalte += 1
        
        spalte += 1
        ws.cell(zeile, spalte).value = worksheet_overview[datum]['accuracy']
        
        zeile += 1
    
    
    
    for datum in sorted(list_of_errors.keys()):
        ws2 = wb.create_sheet('Errors_' + datum, -1)
        
        ws2.cell(1,1).value = 'Startintervall'
        ws2.cell(1,2).value = 'Endintervall'
        ws2.cell(1,3).value = 'Dauer [Intervalle]'
        ws2.cell(1,4).value = label_ai
        ws2.cell(1,5).value = label_real
        
        j = 7
        for h in range(len(behaviors)):
            ws2.cell(1, j+h+1).value = behaviors[h]
            ws2.cell(h+2, j).value = behaviors[h]
            for i in range(len(behaviors)):
                ws2.cell(2+i, j+h+1).value = conf_mat[datum][i, h]
        
        zeile = 2
        for false_interval in list_of_errors[datum]:
            start = int(false_interval['start'])
            end = int(false_interval['end'])
            duration = end - start + 1
            
            if duration <= sensitivity:
                continue
            
            ws2.cell(zeile, 1).value = start
            ws2.cell(zeile, 2).value = end
            ws2.cell(zeile, 3).value = duration
            ws2.cell(zeile, 4).value = behaviors[int(false_interval['ai'])]
            ws2.cell(zeile, 5).value = behaviors[int(false_interval['real'])]
            
            zeile += 1
            
        
    
    wb.save(out)




def read_worksheet(wb, wsname):
    alphabet = list(string.ascii_uppercase)
    sheet = wb[wsname]
    rows = sheet.rows
    first_row = [cell.value for cell in next(rows)]
    data = {}
    z = 2
    for row in rows:
        record = {}
        j = 0
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[alphabet[j]] = cell.value.strip()
            else:
                record[alphabet[j]] = cell.value
            j += 1
        data[z] = record
        z += 1
    return data

def _read_inputfile(inputfile, behavior_mapping):
        
    def _read_csv(filepath, behavior_mapping = behavior_mapping):
        """ 
            Requires path to csv name || start || end || standing || lying || sleeping || out 
            Outputs [ 1,2,1,1,1,2,1,1,1,... ]
        """
        
        ret = []
        
        with open(filepath, "r") as f:

            csv_real = csv.reader(f, delimiter=',')
            j = 0                
            for row in csv_real:
                
                j += 1
                if j == 1:
                    continue
                ret.append( behavior_mapping[np.argmax([ float(row[3]), float(row[4]), float(row[5]), float(row[6]) ])] )
        return ret

    def _read_xlsx(xlsx_file, behavior_mapping = behavior_mapping):
        def get_sequence(wb):
            
            ws = read_worksheet(wb, 'Zeitintervalle')
            ret = []
            
            j = 2
            while j <= max(ws.keys()):
                ret.append( behavior_mapping[int(ws[j]['F'])] )
                j += 1
            
            return ret
        
        data_wb =  load_workbook(xlsx_file)                
        sequence = get_sequence(data_wb)                
        data_wb.close()
        
        return sequence


    if not os.path.exists(inputfile):
        print('ERROR: File does not exist', inputfile)
        return []
    
    if inputfile.endswith('.csv'):
        ret = _read_csv(inputfile)
    elif inputfile.endswith('.xlsx'):
        ret = _read_xlsx(inputfile)
    return ret



def _evaluate_performance(boris_csvs, xlsx, out, behaviors, behavior_mapping, 
                          comparision_mode, individual_code, dates, mode):
    if len(boris_csvs) != len(xlsx):
        print("Error: Input sizes do not match.")
        return
    
    
        
    worksheet_overview = {}
    
    anzahlen_ai = [ [] for j in behaviors ]
    anzahlen_vid = [ [] for j in behaviors ]
    f_scores = [ [] for j in behaviors ]
    dauern_ai = [ [] for j in behaviors ]
    dauern_vid = [ [] for j in behaviors ]
    med_len_ai = [ [] for j in behaviors ]
    med_len_vid = [ [] for j in behaviors ]
    
    accuracies = []
    error_sequences = {}
    conf_mats = {}
    
    for j in range(len(boris_csvs)):
        
        datum = dates[j]
        boris_file = boris_csvs[j]
        xlsx_file = xlsx[j]
        
        print(datum)
        vid_s, vid_e = get_vid_start_end(individual_code)        
        pp_ruleset = get_pp_rule(individual_code, mode, True)
        
        if comparision_mode == 'realPP-ai':
            
            if not os.path.exists(xlsx_file):
                print("ERROR: File not found.", xlsx_file)
                continue
            if not os.path.exists(boris_file):
                print("ERROR: File not found.", boris_file)
                continue
            
            real_sequence = _read_inputfile(boris_file, behavior_mapping)
            ai_sequence = _read_inputfile(xlsx_file, behavior_mapping)
        
            real_sequence = postprocess_night(input_sequence = real_sequence, 
                           behavior_mapping = behavior_mapping,
                           post_processing_rules = pp_ruleset,
                           current_video_start = vid_s,
                           current_video_end = vid_e,
                           apply_postprocessing = True,
                           add_observation_time = True)
            
            ai_sequence = postprocess_night(input_sequence = ai_sequence, 
                           behavior_mapping = behavior_mapping,
                           post_processing_rules = pp_ruleset,
                           current_video_start = vid_s,
                           current_video_end = vid_e,
                           apply_postprocessing = False,
                           add_observation_time = False)
            
            label_real = 'realPP'
            label_ai = 'AI'
        
        if comparision_mode == 'real-realPP':
            
            
            if not os.path.exists(boris_file):
                print("ERROR: File not found.", boris_file)
                continue
            
            real_sequence = _read_inputfile(boris_file, behavior_mapping)
            ai_sequence = _read_inputfile(boris_file, behavior_mapping)
            
        
            real_sequence = postprocess_night(input_sequence = real_sequence, 
                           behavior_mapping = behavior_mapping,
                           post_processing_rules = pp_ruleset,
                           current_video_start = vid_s,
                           current_video_end = vid_e,
                           apply_postprocessing = False,
                           add_observation_time = True)
            
            ai_sequence = postprocess_night(input_sequence = ai_sequence, 
                           behavior_mapping = behavior_mapping,
                           post_processing_rules = pp_ruleset,
                           current_video_start = vid_s,
                           current_video_end = vid_e,
                           apply_postprocessing = True,
                           add_observation_time = True)
            
            label_real = 'real'
            label_ai = 'realPP'
            
        if comparision_mode == 'real-ai':
            
            if not os.path.exists(xlsx_file):
                print("ERROR: File not found.", xlsx_file)
                continue
            if not os.path.exists(boris_file):
                print("ERROR: File not found.", boris_file)
                continue
            
            real_sequence = _read_inputfile(boris_file, behavior_mapping)
            ai_sequence = _read_inputfile(xlsx_file, behavior_mapping)

            real_sequence = postprocess_night(input_sequence = real_sequence, 
                           behavior_mapping = behavior_mapping,
                           post_processing_rules = pp_ruleset,
                           current_video_start = vid_s,
                           current_video_end = vid_e,
                           apply_postprocessing = False,
                           add_observation_time = True)
            
            ai_sequence = postprocess_night(input_sequence = ai_sequence, 
                           behavior_mapping = behavior_mapping,
                           post_processing_rules = pp_ruleset,
                           current_video_start = vid_s,
                           current_video_end = vid_e,
                           apply_postprocessing = False,
                           add_observation_time = False)
            
            label_real = 'real'
            label_ai = 'AI'
            
            
        
        anzahl_vid, anzahl_ai, dauer_vid, dauer_ai, median_orig, median_ai, acc, f_score = compare_one_night( real_sequence, ai_sequence )
        
        worksheet_overview[datum] = {}
        for j in range(len(behaviors)):
            worksheet_overview[datum]['Anzahl_Phasen_'+behaviors[j]] = [int(anzahl_vid[j]), int(anzahl_ai[j])]
            anzahlen_ai[j].append( int(anzahl_ai[j]) )
            anzahlen_vid[j].append( int(anzahl_vid[j]) )
            
            worksheet_overview[datum]['Anteil_'+behaviors[j]] = [np.round(float(dauer_vid[j]),2), np.round(float(dauer_ai[j]),2)]
            dauern_ai[j].append( np.round(dauer_ai[j], 4) )
            dauern_vid[j].append( np.round(dauer_vid[j], 4) )
            
            worksheet_overview[datum]['Median_Length_'+behaviors[j]] = [np.round(float(median_orig[j]),2), np.round(float(median_ai[j]),2)]
            med_len_ai[j].append( int(median_ai[j]) )
            med_len_vid[j].append( int(median_orig[j]) )
            
            worksheet_overview[datum]['fscore_'+behaviors[j]] = np.round(f_score[j], 3)
            f_scores[j].append(np.round(f_score[j], 3))
        worksheet_overview[datum]['accuracy'] = np.round(acc, 3)
        accuracies.append(np.round(acc, 3))
        
        
        
        # get error sequences
        error_seq, cm = describe_errors( real_sequence, ai_sequence )
        error_sequences[datum] = error_seq
        conf_mats[datum] = cm
        
        
        
        
    worksheet_overview['AVG'] = {}
    worksheet_overview['SEM'] = {}
    
    for j in range(len(behaviors)):
        worksheet_overview['AVG']['Anzahl_Phasen_'+behaviors[j]] = [np.mean(anzahlen_vid[j]), np.mean(anzahlen_ai[j])]
        worksheet_overview['SEM']['Anzahl_Phasen_'+behaviors[j]] = [ StandardError(anzahlen_vid[j]), StandardError(anzahlen_ai[j])]
        
        worksheet_overview['AVG']['Anteil_'+behaviors[j]] = [np.mean(dauern_vid[j]), np.mean(dauern_ai[j])]
        worksheet_overview['SEM']['Anteil_'+behaviors[j]] = [ StandardError(dauern_vid[j]), StandardError(dauern_ai[j])]
        
        worksheet_overview['AVG']['Median_Length_'+behaviors[j]] = [np.mean(med_len_vid[j]), np.mean(med_len_ai[j])]
        worksheet_overview['SEM']['Median_Length_'+behaviors[j]] = [ StandardError(med_len_vid[j]), StandardError(med_len_ai[j])]
        
        worksheet_overview['AVG']['fscore_'+behaviors[j]] = np.mean(f_scores[j])
        worksheet_overview['SEM']['fscore_'+behaviors[j]] = StandardError(f_scores[j])
    
    worksheet_overview['AVG']['accuracy'] = np.mean(accuracies)
    worksheet_overview['SEM']['accuracy'] = StandardError(accuracies)
    
    write_outputfile(worksheet_overview, list_of_errors = error_sequences, conf_mat = conf_mats, 
                     out = out, behaviors = behaviors, sensitivity = ERROR_TOLERANCE_INTERVALS,
                     label_real = label_real, label_ai = label_ai)
    

    
def generate_accuracy(compare_real_ai = COMPARE_REAL_WITH_AI,
                      compare_real_pp = COMPARE_REAL_WITH_REALPP,
                      compare_realpp_ai = COMPARE_REALPP_WITH_AI,
                      total = TOTAL,
                      binary = BINARY,
                      individuals = INDIVIDUAL,
                      spec = SPECIES,
                      zoo = ZOO,
                      behaviors = BEHAVIORS):

    
    comparisions = [compare_real_ai, compare_real_pp, compare_realpp_ai]
    comparision_codes = ['real-ai', 'real-realPP', 'realPP-ai']
    modes = [total, binary]
    mode_codes = ['total', 'binary']
    
    for j in range(len(individuals)):
        individual_code = '{}_{}_{}'.format(spec[j], zoo[j], individuals[j])
        print("******************************************************")
        print('Individual: {}'.format(individual_code))
        
        for i in range(len(comparisions)):
            if not comparisions[i]:
                continue
            for m in range(len(modes)):
                if not modes[m]:
                    continue
                
                comparision = comparision_codes[i]
                mode = mode_codes[m]

                output_xlsx = get_output_xlsx(mode, comparision, j)
                csv_folder, xlsx_folder = get_input_folders(mode, j)
                dates_to_evaluate = get_dates_to_evaluate(csv_folder, j)
                print('Mode: {}, Comparision: {}'.format(mode, comparision))
                
                behavior_mapping, xlsx_ext = get_behavior_mapping_and_extensions(mode)
                
                boris_csvs = [ csv_folder + date + '_' + individual_code + '_SUM-7s_pred.csv' for date in sorted(dates_to_evaluate)]
                ai_xlsx = [ '{}{}_{}_statistics{}.xlsx'.format(xlsx_folder, date, individual_code, xlsx_ext) for date in sorted(dates_to_evaluate)] 
                
                _evaluate_performance(boris_csvs = boris_csvs, 
                                      xlsx = ai_xlsx, 
                                      out = output_xlsx, 
                                      behaviors = behaviors,
                                      behavior_mapping = behavior_mapping,
                                      comparision_mode = comparision,
                                      individual_code = individual_code,
                                      dates = dates_to_evaluate,
                                      mode = mode)


#############################################################################
# Code for generating the overview file 
#############################################################################

# Input structure: Elen_Kronberg_1/Elen_Kronberg_1_Accuracy_{}-{}.xlsx {real-ai, real-realPP, realPP-ai}, {total, binary}

def _get_accuracy_sheet(individualcode, code, mode, basefolder):    
    
    if not os.path.exists(basefolder + individualcode):
        return []
    
    return [ basefolder + individualcode + '/' + f for f in sorted(os.listdir(basefolder + individualcode)) if f.endswith('{}-{}.xlsx'.format(code, mode)) ]
    

def _get_all_accuracy_sheets(basefolder):
    
    comparision_codes = ['real-ai', 'real-realPP', 'realPP-ai']
    modes = ['total', 'binary']
    ret = {}
    
    for mode in modes:
        ret[mode] = {}
        for code in comparision_codes:
            ret[mode][code] = {}
    
            for individualcode in sorted( os.listdir(basefolder) ):
                tmp = _get_accuracy_sheet(individualcode, code, mode, basefolder)
                if len(tmp) == 1:
                    ret[mode][code][individualcode] = tmp[0]
            
    return ret


def calculate_statistics(xlsx_file, conf_spalten = CONF_SPALTEN, conf_zeilen = CONF_ZEILEN):
    ret = { 'Accuracy': 0,
                        'f-Stehen': 0,
                        'f-Liegen': 0,
                        'f-Schlafen': 0,
                        
                        '% Stehen (real avg)': 0,
                        '% Stehen (ki avg)': 0,
                        '% Liegen (real avg)': 0,
                        '% Liegen (ki avg)': 0,
                        '% Schlafen (real avg)': 0,
                        '% Schlafen (ki avg)': 0,
                        
                        'Med_Dauer Stehen (real avg)': 0,
                        'Med_Dauer Stehen (ki avg)': 0,
                        'Med_Dauer Liegen (real avg)': 0,
                        'Med_Dauer Liegen (ki avg)': 0,
                        'Med_Dauer Schlafen (real avg)': 0,
                        'Med_Dauer Schlafen (ki avg)': 0,
                        
                        '# Stehen': 0,
                        'Stehen (Abw)': 0,
                        'max. Abweichung [%] Stehen': 0,
                        '# Liegen': 0,
                        'Liegen (Abw)' : 0,
                        'max. Abweichung [%] Liegen': 0,
                        'Phasen Schlafen (max. Abweichung)': 0,
                        '# Schlafen': 0,
                        'Schlafen (Abw)' : 0,
                        'max. Abweichung [%] Schlafen': 0,
                        'max. Abweichung (Phasen) Out': 0,
                        
                        'Phasen Stehen (real avg)': 0,
                        'Phasen Stehen (ki avg)': 0,
                        'mittlere Abweichung [%] Stehen': 0,
                        'Phasen Liegen (real avg)': 0,
                        'Phasen Liegen (ki avg)': 0,
                        'mittlere Abweichung [%] Liegen': 0,
                        'Phasen Schlafen (real avg)': 0,
                        'Phasen Schlafen (ki avg)': 0,
                        'mittlere Abweichung [%] Schlafen': 0,
                        'mittlere Abweichung (Phasen) Out': 0,
                        
                        '# Nächte': 0,
                        '# Fehler': 0,
                        'Fehler / Nacht': 0,
                        'min. conf. Mat. Stehen': 0,
                        'min. conf. Mat. Liegen': 0,
                        'min. conf. Mat. Schlafen': 0,
                        'min. conf. Mat. Out': 0,
                        'avg. conf. Mat. Stehen': 0,
                        'avg. conf. Mat. Liegen': 0,
                        'avg. conf. Mat. Schlafen': 0,
                        'avg. conf. Mat. Out': 0
                    }
    
    all_sheets = pd.read_excel(xlsx_file, None)
    

    # collecting information from the overview sheet    
    overview = all_sheets['Overview']
    
    l = len(overview.iloc[:, 24])
    ret['Accuracy'] = overview.iloc[l-2, 40] # adjust
    ret['f-Stehen'] = overview.iloc[l-2, 34] # adjust
    ret['f-Liegen'] = overview.iloc[l-2, 35] # adjust
    ret['f-Schlafen'] =  overview.iloc[l-2, 36] # adjust
    ret['# Nächte'] = len(all_sheets) - 1
    
    abweichungen = {'Stehen': [], 'Liegen': [], 'Schlafen': []}
    abw_total = {'Stehen': [], 'Liegen': [], 'Schlafen': []} 
    values = {'Stehen': [], 'Liegen': [], 'Schlafen': []}
    
    zeile = 1
    while zeile <= len(overview.index) - 4:        
        ind = 0   
        for b in abweichungen.keys():        
            real = overview.iloc[zeile, 1 + ind]
            ai = overview.iloc[zeile, 2 + ind]
            ind += 2
            values[b].append(real)
            if real == 0:
                abweichungen[b].append(0)
            elif real > 0:
                abweichungen[b].append( np.round(np.abs( 1 - ai/real),3))
            abw_total[b].append( np.abs(ai-real) )
            
        zeile += 1
    
    ind = 0
    for b in abweichungen.keys():
        max_ind = np.argmax( abweichungen[b] )
        ret['max. Abweichung [%] {}'.format(b)] =  abweichungen[b][max_ind]  
        ret['# {}'.format(b)] = values[b][max_ind]  
        ret['{} (Abw)'.format(b)]  =  abw_total[b][max_ind]       
        real_avg = overview.iloc[ len(overview.index) -2, 1 + ind]
        ai_avg = overview.iloc[ len(overview.index) -2, 2 + ind]        
        avg = 0 if real_avg == 0 else np.round(np.abs( 1 - ai_avg/real_avg), 3)       
        avg_diff = np.round(ai_avg - real_avg, 3)
                
        ret['mittlere Abweichung [%] {}'.format(b)] = avg
        ret['Phasen {} (real avg)'.format(b)] = real_avg
        ret['Phasen {} (ki avg)'.format(b)] = avg_diff
        
        
        # todo: abweichung anteil
        ret['% {} (real avg)'.format(b)] = overview.iloc[ len(overview.index) -2, 12 + ind]
        ret['% {} (ki avg)'.format(b)] = overview.iloc[ len(overview.index) -2, 13 + ind] - overview.iloc[ len(overview.index) -2, 12 + ind]
        
        # todo: abweichung mediane länge
        ret['Med_Dauer {} (real avg)'.format(b)] = overview.iloc[ len(overview.index) -2, 23 + ind]
        ret['Med_Dauer {} (ki avg)'.format(b)] = overview.iloc[ len(overview.index) -2, 24 + ind] - overview.iloc[ len(overview.index) -2, 23 + ind]
        
        ind += 2
        
    out_abw = 0
    zeile = 1
    while zeile <= len(overview.index) - 4:
        real_out = overview.iloc[zeile, 7]
        ai_out = overview.iloc[zeile, 8]        
        out_abw = max(out_abw, np.abs(real_out - ai_out)) 
        zeile += 1
    
    ret['max. Abweichung (Phasen) Out'] = out_abw
    ret['mittlere Abweichung (Phasen) Out'] = np.abs( overview.iloc[len(overview.index) - 2, 8] - overview.iloc[len(overview.index) - 2, 7]  )
    
    # collecting information from the other sheets
    num_errors = 0
    conf = {}
    
    for b in conf_spalten:
        conf[b] = []
        
    for sheetname in all_sheets.keys():
        if sheetname == 'Overview':
            continue
        
        df = all_sheets[sheetname]
        num_errors += len(df[ df['Startintervall'] > 0])        
        
        zeile = 0
        
        for bzeile in conf_zeilen:
            total = 0
            off = 0
            
            for spalte in conf_spalten:
                total += df[spalte][zeile]
                if spalte != bzeile:
                    off += df[spalte][zeile]
                
            if total == 0:
                prec = 1
            else:
                prec = np.round(1 - off/total, 3)
            
            
            conf[bzeile].append(prec)
            zeile += 1
    ret['# Fehler'] = num_errors        
    for b in conf_zeilen:
        ret['min. conf. Mat. {}'.format(b)] = np.min( conf[b] )
        ret['avg. conf. Mat. {}'.format(b)] = np.mean( conf[b] )
    
    ret['Fehler / Nacht'] = np.round(ret['# Fehler']/ret['# Nächte'],2)
    
    return ret
    
def write_worksheet_overview(inputdict, outputname, comparision):
    
    
    data = {}
    modes = ['total', 'binary']   

    
    for mode in modes:
        data[mode] = pd.DataFrame({'Individuum' : [],
                      'Accuracy': [],
                        'f-Stehen': [],
                        'f-Liegen': [],
                        'f-Schlafen': [],
                        
                        
                        '% Stehen (real avg)': [],
                        '% Stehen (ki avg)': [],
                        '% Liegen (real avg)': [],
                        '% Liegen (ki avg)': [],
                        '% Schlafen (real avg)': [],
                        '% Schlafen (ki avg)': [],
                        
                        'Med_Dauer Stehen (real avg)': [],
                        'Med_Dauer Stehen (ki avg)': [],
                        'Med_Dauer Liegen (real avg)': [],
                        'Med_Dauer Liegen (ki avg)': [],
                        'Med_Dauer Schlafen (real avg)': [],
                        'Med_Dauer Schlafen (ki avg)': [],
                        
                        '# Stehen': [],
                        'Stehen (Abw)': [],
                        'max. Abweichung [%] Stehen': [],
                        '# Liegen': [],
                        'Liegen (Abw)' : [],
                        'max. Abweichung [%] Liegen': [],
                        'Phasen Schlafen (max. Abweichung)': [],
                        '# Schlafen': [],
                        'Schlafen (Abw)' : [],
                        'max. Abweichung [%] Schlafen': [],
                        'max. Abweichung (Phasen) Out': [],
                        'Phasen Stehen (real avg)': [],
                        'Phasen Stehen (ki avg)': [],
                        'mittlere Abweichung [%] Stehen': [],
                        'Phasen Liegen (real avg)': [],
                        'Phasen Liegen (ki avg)': [],
                        'mittlere Abweichung [%] Liegen': [],
                        'Phasen Schlafen (real avg)': [],
                        'Phasen Schlafen (ki avg)': [],
                        'mittlere Abweichung [%] Schlafen': [],
                        'mittlere Abweichung (Phasen) Out': [],
                        '# Nächte': [],
                        '# Fehler': [],
                        'Fehler / Nacht': [],
                        'min. conf. Mat. Stehen': [],
                        'min. conf. Mat. Liegen': [],
                        'min. conf. Mat. Schlafen': [],
                        'min. conf. Mat. Out': [],
                        'avg. conf. Mat. Stehen': [],
                        'avg. conf. Mat. Liegen': [],
                        'avg. conf. Mat. Schlafen': [],
                        'avg. conf. Mat. Out': []}
                    )
        
        
    
        for individual_code in inputdict[mode][comparision].keys():   
            
            tmp_dict = {'Individuum' : [],
                      'Accuracy': [],
                        'f-Stehen': [],
                        'f-Liegen': [],
                        'f-Schlafen': [],
                        
                        '% Stehen (real avg)': [],
                        '% Stehen (ki avg)': [],
                        '% Liegen (real avg)': [],
                        '% Liegen (ki avg)': [],
                        '% Schlafen (real avg)': [],
                        '% Schlafen (ki avg)': [],
                        
                        'Med_Dauer Stehen (real avg)': [],
                        'Med_Dauer Stehen (ki avg)': [],
                        'Med_Dauer Liegen (real avg)': [],
                        'Med_Dauer Liegen (ki avg)': [],
                        'Med_Dauer Schlafen (real avg)': [],
                        'Med_Dauer Schlafen (ki avg)': [],
                        
                        '# Stehen': [],
                        'Stehen (Abw)': [],
                        'max. Abweichung [%] Stehen': [],
                        '# Liegen': [],
                        'Liegen (Abw)' : [],
                        'max. Abweichung [%] Liegen': [],
                        'Phasen Schlafen (max. Abweichung)': [],
                        '# Schlafen': [],
                        'Schlafen (Abw)' : [],
                        'max. Abweichung [%] Schlafen': [],
                        'max. Abweichung (Phasen) Out': [],
                        'Phasen Stehen (real avg)': [],
                        'Phasen Stehen (ki avg)': [],
                        'mittlere Abweichung [%] Stehen': [],
                        'Phasen Liegen (real avg)': [],
                        'Phasen Liegen (ki avg)': [],
                        'mittlere Abweichung [%] Liegen': [],
                        'Phasen Schlafen (real avg)': [],
                        'Phasen Schlafen (ki avg)': [],
                        'mittlere Abweichung [%] Schlafen': [],
                        'mittlere Abweichung (Phasen) Out': [],
                        '# Nächte': [],
                        '# Fehler': [],
                        'Fehler / Nacht': [],
                        'min. conf. Mat. Stehen': [],
                        'min. conf. Mat. Liegen': [],
                        'min. conf. Mat. Schlafen': [],
                        'min. conf. Mat. Out': [],
                        'avg. conf. Mat. Stehen': [],
                        'avg. conf. Mat. Liegen': [],
                        'avg. conf. Mat. Schlafen': [],
                        'avg. conf. Mat. Out': []}
            
            tmp_dict['Individuum'].append(individual_code)
            xlsx_file = inputdict[mode][comparision][individual_code]
            inf = calculate_statistics(xlsx_file)
            

            for key in inf.keys():
                tmp_dict[key].append( inf[key] )
            
            data[mode] = data[mode].append( pd.DataFrame(tmp_dict) )

            
    

    with pd.ExcelWriter(outputname) as writer:  
        for mode in modes:
            data[mode].to_excel(writer, sheet_name=mode)


def generate_bovids_files(ausw = CSV_FOLDER, outp = OUTPUT_FOLDER_BOVIDS, individuals = INDIVIDUALCODES_BOVIDS):
    species, zoo, individual = _get_speczoo(individuals)
    
    for mode in ['total', 'binary']:
        print("Mode: ", mode)
        behavior_mapping, extension = get_behavior_mapping_and_extensions(mode)
        for j in range(len(species)):
            csv_input_folder, _ = get_input_folders(mode, j, csv_base = ausw, ai_base = '',  species = species, zoo = zoo, ind = individual)
            dates = get_dates_to_evaluate(csv_input_folder, j, spec = species, zoo = zoo, ind = individual)
            dates = sorted(dates)
            xlsx_output_folder = '{}{}/{}/Auswertung/Boris_KI/Boris_PP/'.format(outp, species[j], zoo[j])
            
            individual_code = '{}_{}_{}'.format(species[j], zoo[j], individual[j])
            vid_s, vid_e = get_vid_start_end(individual_code)        
            pp_ruleset = get_pp_rule(individual_code, mode, True)
                
            print("******************************************************")
            print('Individual: {}'.format(individual_code))

            boris_csvs = [ csv_input_folder + date + '_' + individual_code + '_SUM-7s_pred.csv' for date in dates]
            
            for h in range(len(boris_csvs)):
                
                boris_file = boris_csvs[h]
                date = dates[h]
                
                if not os.path.exists(boris_file):
                    continue
                   

                real_sequence = _read_inputfile(boris_file, behavior_mapping)            
                real_sequence = postprocess_night(input_sequence = real_sequence, 
                               behavior_mapping = behavior_mapping,
                               post_processing_rules = pp_ruleset,
                               current_video_start = vid_s,
                               current_video_end = vid_e,
                               apply_postprocessing = True,
                               add_observation_time = True)
                
                write_output_xlsx_file(ensemble_sparse = real_sequence, 
                                       outputfolder = xlsx_output_folder, 
                                       datum = date, 
                                       individual_code = individual_code, 
                                       mode = mode)
            

def generate_overview(inp = INPUT_FOLDER_FOR_OVERVIEW, outp = OUTPUT_FOLDER_OVERVIEW ):
    accuracy_dict = _get_all_accuracy_sheets(inp)
    comparision_codes = ['real-ai', 'real-realPP', 'realPP-ai']
    for comp in comparision_codes:
        outpf = outp + 'overview_'+comp+'.xlsx'
        ensure_directory(outpf)
        write_worksheet_overview(accuracy_dict, outpf, comp)