#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "Max Hahn-Klimroth"
__copyright__ = "Copyright 2021, M. Hahn-Klimroth, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes"]
__license__ = "GPL-3.0"
__version__ = "1.0"
__status__ = "Development"

import os
from collections import Counter
from openpyxl import Workbook

MAX_INTERVALS = 7200
SIZE_PACKAGES = int(3600/7) + 1
MAX_PER_INTERVAL = 4
XLSX_OUTPUT = ''


def get_number_boxes(filename):
    x = sum(1 for line in open(filename) if line.rstrip())
    return x

def get_sequence_per_night(input_folder, 
                           max_intervals, 
                           packages,
                           max_per_interval):
    files = sorted([ f for f in os.listdir(input_folder) if f.endswith('.txt') ])
    amount_images = []
    
    j = 1
    
    while j <= max_intervals:
        interval = str(j)
        cor_file = interval.zfill(7) + '.txt'
        if  cor_file in files:
            num = get_number_boxes( input_folder + cor_file )
        else:
            num = 0
        
        amount_images.append(num)
        j += 1
    
    c = Counter(amount_images)
    
    total_nums = [c[j] for j in [0,1,2,3,4]]
    density = round(sum(amount_images) / (max_intervals*max_per_interval), 3)
    
    subsequences = [amount_images[i:i + packages] for i in range(0, len(amount_images), packages)]
    
    
    subsequence_nums = []
    subsequence_density = []
    
    for seq in subsequences:
        c = Counter(seq)
        subsequence_nums.append([c[j] for j in [0,1,2,3,4]])
        subsequence_density.append(round(sum(seq) / (len(seq)*max_per_interval), 3))
    
    return total_nums, density, subsequence_nums, subsequence_density



def write_xlsx_overview(sequences, xlsx_output, individual_code):
    wb = Workbook()
    ws_standard = wb.active

    print("Preparing " + xlsx_output)
    ws = wb.create_sheet(title=individual_code)
    
    possible_dates = [ date for date in sorted(sequences.keys()) ]
    num_seq_list = [ len(sequences[date][3]) for date in possible_dates]
    amount_sequences = max(num_seq_list)
    
    ws.cell(3, 1).value = 'Übersicht'
    
    ws.cell(4, 1).value = 'Anzahl 0'
    ws.cell(5, 1).value = 'Anzahl 1'
    ws.cell(6, 1).value = 'Anzahl 2'
    ws.cell(7, 1).value = 'Anzahl 3'
    ws.cell(8, 1).value = 'Anzahl 4'
    ws.cell(9, 1).value = 'Dichte'
    
    zeile = 10
    j = 1
    for pack in range(amount_sequences):
        
        
        ws.cell(zeile + 1, 1).value  = 'Sequenz ' + str(j)
        ws.cell(zeile + 2, 1).value = 'Anzahl 0'
        ws.cell(zeile + 3, 1).value = 'Anzahl 1'
        ws.cell(zeile + 4, 1).value = 'Anzahl 2'
        ws.cell(zeile + 5, 1).value = 'Anzahl 3'
        ws.cell(zeile + 6, 1).value = 'Anzahl 4'
        ws.cell(zeile + 7, 1).value = 'Dichte'
        
        j += 1
        zeile += 8
    
    spalte = 2
    num_date = 0
    for datum in possible_dates:
        ws.cell(1, spalte).value = individual_code
        ws.cell(2, spalte).value = datum
        ws.cell(4, spalte).value = sequences[datum][0][0]
        ws.cell(5, spalte).value = sequences[datum][0][1]
        ws.cell(6, spalte).value = sequences[datum][0][2]
        ws.cell(7, spalte).value = sequences[datum][0][3]
        ws.cell(8, spalte).value = sequences[datum][0][4]
        ws.cell(9, spalte).value = sequences[datum][1]
        
        zeile = 10
        j = 0
        for pack in range(num_seq_list[num_date]):
            
            
            ws.cell(zeile + 2, spalte).value = sequences[datum][2][j][0]
            ws.cell(zeile + 3, spalte).value = sequences[datum][2][j][1]
            ws.cell(zeile + 4, spalte).value = sequences[datum][2][j][2] 
            ws.cell(zeile + 5, spalte).value = sequences[datum][2][j][3] 
            ws.cell(zeile + 6, spalte).value = sequences[datum][2][j][4]
            ws.cell(zeile + 7, spalte).value = sequences[datum][3][j]
            
            j += 1
            zeile += 8
        spalte += 1
        num_date += 1
    wb.remove(ws_standard)
    wb.save(xlsx_output)
            
        
    