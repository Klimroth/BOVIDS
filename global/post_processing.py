#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = "Max Hahn-Klimroth, Tobias Kapetanopoulos"
__copyright__ = "Copyright 2020, M. Hahn-Klimroth, T. Kapetanopoulos, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes"]
__license__ = "MIT"
__version__ = "1.1"
__maintainer__ = "M. Hahn-Klimroth"
__status__ = "Development"

"""
Version 2022-01-13: 
    apply_rolling_average, calculate_joint_prediction,
    _single_frame_prediction_to_intervalprediction: added rounding to 6 decimals
"""
import configuration as cf
import copy
import csv, os
import numpy as np
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.collections import PolyCollection
import cv2

def ensure_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)

# just for testing purposes
def _test_night(individual_code, datum, inputpath, outputpath, name = '_SUM-7s_pred.csv'):
    csvfile = inputpath + datum + '_' + individual_code + name
    post_process_night(single_frame_csv = csvfile, position_files = '', 
                       joint_interval_csv = csvfile, 
                       individual_code = individual_code, 
                       datum = datum, is_test = True, 
                       output_folder_prediction = outputpath)

def post_process_night(single_frame_csv, joint_interval_csv, 
                       individual_code, datum, position_files,
                       output_folder_prediction = cf.FINAL_STORAGE_PREDICTION_FILES, 
                       behavior_mapping = cf.BEHAVIOR_MAPPING,
                       is_test = False, 
                       extension = '',
                       post_processing_rules = cf.POST_PROCESSING_RULES['Standard'],
                       truncation_rules = {},
                       truncation_areas = {},
                       current_video_start = 17,
                       current_video_end = 7,
                       behavior_names = cf.BEHAVIOR_NAMES,
                       logfile = '',
                       out_regulations = cf.OUT_REGULATIONS):
    
    
    def _map_behaviors( phase_list, behavior_mapping = behavior_mapping, pp_rules = post_processing_rules):
        """
        Maps phases, e.g. if we want to join lying and sleeping.
        """
        for phase in phase_list:
            phase[1] = behavior_mapping[phase[1]]        
        return phase_list
    
    def _apply_rolling_average( list_of_dist, amount_back, weights, potency = 1.0, behavior_names = behavior_names):
        rolling_avg_per_img = []
        for j in range(len(list_of_dist)):
            go_back_here = min(j, amount_back)        
            curr_dist = np.array([0.0]*len(behavior_names))
            
            for k in range(j - go_back_here, j+1):
                curr_dist += np.around(list_of_dist[k]*weights, 8) #*(potency**(j-k))
            curr_dist = curr_dist / np.linalg.norm(curr_dist, ord=1)
            rolling_avg_per_img.append(np.around(curr_dist, 6))
        
        return rolling_avg_per_img
    
    def _read_csv(filepath):
        """ 
            Requires path to csv name || start || end || standing || lying || sleeping || out 
            Outputs [ [standing, lying, sleeping, out, truncated] ]
        """
        
        ret = []
        
        with open(filepath, "r") as f:

            csv_real = csv.reader(f, delimiter=',')
            j = 0    
            for row in csv_real:
                j += 1
                if j == 1:
                    continue
                ret.append( [ np.float(row[3]), np.float(row[4]), np.float(row[5]), np.float(row[6]), 0.0 ] )
        
        return ret   
    
    def _single_frame_prediction_to_intervalprediction(dist_list, n = cf.INTERVAL_LENGTH, behavior_names = behavior_names):
        list_of_intervals = [dist_list[i * n:(i + 1) * n] for i in range((len(dist_list) + n - 1) // n )]
        ret = []
        for interval in list_of_intervals:
            curr = np.array( [0.0]*len(behavior_names) )
            for img_dist in interval:
                curr += np.around(np.array( img_dist ), 8)
            curr = curr / np.linalg.norm(curr, ord=1)
            
            ret.append(np.around(curr,6))
        return ret
    
    def _calculate_joint_prediction( dist1, dist2, weights, num_behav = 5, is_test = is_test ):
        """
        
    
        Parameters
        ----------
        dist 1: joint images
        dist 2: single frames
        
    
        Returns
        -------
        
            Output: List of intervals, for each 
    
        """
        
        if is_test:
            print("Testing mode on real data.")
            return dist1
        
        len1 = len(dist1)
        len2 = len(dist2)
        
        ret = []
        
        if len1 != len2:
            print("Warning: Different lenghts of predictions.", len1, len2)
        
        minlen = min(len1, len2)
        
        for j in range(minlen):
            pred1 = dist1[j]
            pred2 = dist2[j]
            
            pred = np.around(weights[0] * weights[1]*np.array(pred2), 6)
            
            ret.append(pred)
        
        return ret
    
    def _write_prediction_csv(interval_dist, output_folder, filename, behavior_names = behavior_names, int_len = cf.INTERVAL_LENGTH):
        file_path = output_folder + filename
        ensure_dir(output_folder)
        
        with open(file_path, mode='w+') as csv_out:
            csv_write = csv.writer(csv_out, delimiter=",")
            csv_write.writerow( ["Interval", "Startframe", "Endframe", 
                                 behavior_names[0],
                                 behavior_names[1],
                                 behavior_names[2],
                                 behavior_names[3],
                                 behavior_names[4]] )
            
            h = 1
            for behav_int in interval_dist:
                sf = (h-1)*int_len + 1
                ef = h*int_len
                row = [h, sf, ef, behav_int[0], behav_int[1], behav_int[2] ,behav_int[3], behav_int[4]]
                csv_write.writerow(row)
                h += 1
                
    def _sparse_encoding(interval_dist):
        ret = []
        for interval in interval_dist:
            ret.append( np.argmax( interval ) )
        
        return ret
    
    def _extract_single_phases(np_array, timeinterval = cf.INTERVAL_LENGTH):
        """
        
    
        Parameters
        ----------
        np_array : TYPE
            DESCRIPTION.
        timeinterval : TYPE, optional
            DESCRIPTION. The default is cf.INTERVAL_LENGTH.
    
        Returns
        -------
        phases : [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ]
    
        """
        phases = []
        LastBehav = np_array[0]
        iCurrLen = 1
        j = 1
        start_interval = 1

        for DistLine in np_array:
            
            if LastBehav == DistLine:
                iCurrLen += 1
            else:
                phases.append([iCurrLen*timeinterval, LastBehav, start_interval, j-1])
                iCurrLen = 1
                start_interval = j
                LastBehav = DistLine
            j += 1
        
        if len(phases) == 0:
            phases.append( [iCurrLen*timeinterval, LastBehav, start_interval, j-1] )
        else:
            phases.append([iCurrLen*timeinterval, LastBehav, phases[-1][3] + 1, phases[-1][3] + 1 + iCurrLen])
        return phases
    
    
    def _join_phases(phase_list):
        """        
            Joins consecutive phases of the same behavior to a single phase.

        """
    
        ret = [phase_list[0]]
        
        for j in range(1, len(phase_list)):
            if phase_list[j][1] == phase_list[j-1][1]:
                ret[-1][0] += phase_list[j][0]
                ret[-1][3] = phase_list[j][3]
            else:
                ret.append(phase_list[j])
        
        return ret
    
    
    def _remove_short_phases(phase_list, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Removes those short phases which are very unlikely due to configuration
        """
        
        def _check_first_phase(phase_list):
            
            if len(phase_list) <= 1:
                return False
            
            if phase_list[0][1] == 4:
                phase_list[0][1] = 3
                return True
            l = phase_list[0][0]
            if l < 5*9*interval_len:
                phase_list[0][1] = phase_list[1][1]
                return True
            
            return False
            
        phase_list = _join_phases(phase_list)
        
        if len(phase_list) < 3:
            return phase_list
        
        
        changes_done = True
        
        while changes_done:
            
            last_behavior = phase_list[0][1]
            current_behavior = None
            next_behavior = None
            
            changes_done = False
            changes_done = _check_first_phase(phase_list)
            
            for j in range(1, len(phase_list)-1):
                last_behavior = phase_list[j-1][1]
                current_behavior = phase_list[j][1]
                next_behavior = phase_list[j+1][1]
                
                # lying
                if [last_behavior, current_behavior, next_behavior] == [2,1,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SLS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,0]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_ALA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_ALS']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,1,0]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SLA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,1,0]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OLA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_ALO']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,1,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OLS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,1,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SLO']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,1,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OLO']*interval_len:
                        phase_list[j][1] = 3
                        changes_done = True
                        
                        
                # standing
                elif [last_behavior, current_behavior, next_behavior] == [2,0,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SAS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LAL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_LAS']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,0,1]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SAL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,0,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_SAO']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,0,1]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OAL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_LAO']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,0,2]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OAS']*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,0,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OAO']*interval_len:
                        phase_list[j][1] = 3
                        changes_done = True
                        
                        
                        
                # sleeping
                elif [last_behavior, current_behavior, next_behavior] == [0,2,0]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_ASA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,2,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_ASL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LSL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,0]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LSA']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True                
                elif [last_behavior, current_behavior, next_behavior] == [0,2,3]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_ASO']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,2,1]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_OSL']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,2,0]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_OSA']*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,3]: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_LSO']*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [3,2,3]:
                    if phase_list[j][0] < pp_rules['MIN_LEN_OSO']*interval_len:
                        phase_list[j][1] = 3
                        changes_done = True
                

                        
                # truncation
                elif current_behavior == 4: 
                    if phase_list[j][0] < pp_rules['MIN_LEN_TRUNCATION']*interval_len:
                        phase_list[j][1] = last_behavior
                        changes_done = True     
                    elif phase_list[j][0] >= pp_rules['MIN_LEN_TRUNCATION_SWAP']*interval_len:
                        phase_list[j][1] = pp_rules['TRUNCATION_REAL_BEHAVIOR_LONG']
                        changes_done = True 
                    elif phase_list[j][0] >= pp_rules['MIN_LEN_TRUNCATION']*interval_len and phase_list[j][0] < pp_rules['MIN_LEN_TRUNCATION_SWAP']*interval_len:
                        phase_list[j][1] = pp_rules['TRUNCATION_INTERMEDIATE']
            
            phase_list = _join_phases(phase_list)
            phase_list, x = _remove_out_fluctuation(phase_list) 

            if changes_done or x:
                changes_done = True
            else:
                changes_done = False
            
        return phase_list

    def _remove_out(phase_list, logfile, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """
        Removes those phases of behavior out which are shorter than the value given by the configuration
        """
        j = 0
        for phase in phase_list:
            if phase[1] == 3 and j >= 1: # it is out
                if phase[0] <= pp_rules['MIN_LEN_OUT']*interval_len and phase[2] > pp_rules['MIN_TIME_OUT']:
                    phase[1] = phase_list[j-1][1]
                    with open(logfile, 'a+') as file:
                        file.writelines(['{} {} {}\n'.format(phase[0], phase[1], phase[2])])
            j += 1
        
        return phase_list

    
    def _remove_truncated_images(phase_list, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """
        Removes those phases of truncation out which are longer than the value given by the configuration
        """
        j = 0
        for phase in phase_list:
            if phase[1] == 4 and j >= 1: # it is truncated
                if phase[0] >= pp_rules['MIN_LEN_TRUNCATION_SWAP']*interval_len:
                    phase[1] = pp_rules['TRUNCATION_REAL_BEHAVIOR_LONG']
                elif phase[0] >= pp_rules['MIN_LEN_TRUNCATION']*interval_len and phase[0] < pp_rules['MIN_LEN_TRUNCATION_SWAP']*interval_len:
                    phase[1] = pp_rules['TRUNCATION_INTERMEDIATE']
            j += 1
        
        return phase_list
    
    
    def _get_time(interval_num, video_start = cf.GLOBAL_STARTING_TIME, interval_len = cf.INTERVAL_LENGTH):
        """
        
    
        Parameters
        ----------
        interval_num : TYPE
            DESCRIPTION.
        video_start : TYPE, optional
            DESCRIPTION. The default is GLOBAL_STARTING_TIME (when the standard observation should start).
    
        Returns
        -------
        startzeit, endzeit, startframe, endframe
    
        """
        startframe = (interval_num-1)*interval_len + 1
        endframe = interval_num*interval_len
        
        start_hours = startframe // (60*60)
        start_minutes = (startframe % (60*60)) // 60
        start_seconds = startframe % 60
        
        end_hours = endframe // (60*60)
        end_minutes = (endframe % (60*60)) // 60
        end_seconds = endframe % 60
        
        start_time = str((video_start + start_hours)%24).zfill(2) + ":" + str(start_minutes).zfill(2) + ":" + str(start_seconds).zfill(2)
        end_time = str((video_start + end_hours)%24).zfill(2) + ":" + str(end_minutes).zfill(2) + ":" + str(end_seconds).zfill(2)
        
        return start_time, end_time, startframe, endframe
    
    def _get_sparse_intervals_from_phases(phases):
       
        new_intervallist = [] # sequence of behaviors
        for j in range(1, len(phases)+1):
            curr_phase = phases[j-1]
            curr_start = curr_phase[2]
            curr_end = curr_phase[3]
            curr_dur = curr_end - curr_start
            for i in range(curr_dur + 1):
                new_intervallist.append(curr_phase[1])
        return new_intervallist
     
    
    def _write_xlsx_statistics(phases_ordered, phases_unordered, interval_list, outputfolder, 
                               filename, names_behav = cf.BEHAVIOR_NAMES, extension=extension):
        wb = Workbook()
        ws = wb.active
        ws.title = "Zeitintervalle"
        
        ws["A1"] = "Zeitintervall"
        ws["B1"] = "Startzeit"
        ws["C1"] = "Endzeit"
        ws["D1"] = "Startframe"
        ws["E1"] = "Endframe"
        ws["F1"] = "Verhaltenscode"
        ws["G1"] = "Verhaltenswort"
        
        
        for j in range(len(interval_list)):
            starttime, endtime, startframe, endframe = _get_time(j+1)
            ws["A"+str(j+2)] = j+1
            ws["B"+str(j+2)] = starttime
            ws["C"+str(j+2)] = endtime
            ws["D"+str(j+2)] = startframe
            ws["E"+str(j+2)] = endframe
            ws["F"+str(j+2)] = interval_list[j]
            ws["G"+str(j+2)] = names_behav[interval_list[j]]
            
        
                    
        #  [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ]
        # phases: [ [dauer (sek), verhaltenscode, startinterval, endinterval] ] (stat sheet 2)
        ws2 = wb.create_sheet("Aktivitätsphasen_geordnet", -1)
        
        ws2["A1"] = "Phase"
        ws2["B1"] = "Nummer der Phase des Verhaltens"
        ws2["C1"] = "Startzeit"
        ws2["D1"] = "Endzeit"
        ws2["E1"] = "Startintervall"
        ws2["F1"] = "Endintervall"
        ws2["G1"] = "Länge [sec]"
        ws2["H1"] = "Verhaltenscode"
        ws2["I1"] = "Verhaltenswort"
        
        j = 0
        amount_phases = [0]*len(names_behav)
        for phase in phases_unordered:
            amount_phases[phase[1]] += 1
            starttime, _, _, _ = _get_time(phase[2])
            _, endtime, _, _ = _get_time(phase[3])
            
            ws2["A"+str(j+2)] = j+1
            ws2["B"+str(j+2)] = amount_phases[phase[1]] 
            ws2["C"+str(j+2)] = starttime
            ws2["D"+str(j+2)] = endtime
            ws2["E"+str(j+2)] = phase[2]
            ws2["F"+str(j+2)] = phase[3]
            ws2["G"+str(j+2)] = phase[0]
            ws2["H"+str(j+2)] = phase[1]
            ws2["I"+str(j+2)] = names_behav[phase[1]]
            
            j += 1
            
            
        # phases_ordered [ [ [dauer (s), 0, startinterval, endinterval] ... ], [dauer (s), 1, startinterval, endinterval], ... ] (stat sheet 3) 
        ws3 = wb.create_sheet("Aktivitätsphasen_Übersicht", -1)
        colname=["A", "B", "C", "D", "E", "F"]
        
        for i in range(4):
            ws3[colname[i]+"1"] = names_behav[i]
            j = 0
            for phase_stand in phases_ordered[i]:
                ws3[colname[i]+str(j+2)] = phase_stand
                j += 1
        
        
        
        
        ws4 = wb.create_sheet("Statistik", -1)
        
        ws4["A2"] = "Anzahl Phasen"
        ws4["A3"] = "Gesamtdauer [sec]"
        ws4["A4"] = "Anteil [%]"
        
        ws4["A6"] = "Median [sec]"
        ws4["A7"] = "0.25-Quantil"
        ws4["A8"] = "0.75-Quantil"
        ws4["A10"] = "Mean [sec]"
        ws4["A11"] = "SEM"
        
        total_duration = 0
        for j in range(len(names_behav)):
            total_duration += np.sum(phases_ordered[j])
        for j in range(len(names_behav)):
            amount = len(phases_ordered[j])
            ws4[colname[j+1]+"1"] = names_behav[j]
            if amount > 0:
                ws4[colname[j+1]+"2"] = amount
                ws4[colname[j+1]+"3"] = np.sum(phases_ordered[j])
                ws4[colname[j+1]+"4"] = round(100 * np.sum(phases_ordered[j]) / total_duration, 1)
                
                ws4[colname[j+1]+"6"] = np.median(phases_ordered[j])
                ws4[colname[j+1]+"7"] = np.quantile(phases_ordered[j], 0.25)
                ws4[colname[j+1]+"8"] = np.quantile(phases_ordered[j], 0.75)
                ws4[colname[j+1]+"10"] = np.mean(phases_ordered[j])
                ws4[colname[j+1]+"11"] = np.std(phases_ordered[j]) / np.sqrt(amount)
            else:
                ws4[colname[j+1]+"2"] = amount
                ws4[colname[j+1]+"3"] = 0
                ws4[colname[j+1]+"4"] = 0
                
                ws4[colname[j+1]+"6"] = 0
                ws4[colname[j+1]+"7"] = 0
                ws4[colname[j+1]+"8"] = 0
                ws4[colname[j+1]+"10"] = 0
                ws4[colname[j+1]+"11"] = 0
        
        ensure_dir(outputfolder)
        wb.save(outputfolder + filename + extension + '.xlsx')
    
    def _draw_timeline(phases_unordered, save_path, title, behav_names2 = cf.BEHAVIOR_NAMES, 
                       colormapping = cf.COLOR_MAPPING, extension = extension):
        
        behav_names = copy.deepcopy(behav_names2)
        data = []
        for phase in phases_unordered:
            startframe = phase[2]
            endframe = phase[3]
            behavior = behav_names[phase[1]]
            
            data.append((startframe, endframe, behavior))
        
        cats ={}
        for j in range(len(behav_names)):
            cats[behav_names[j]] = (j+1)/2
        
        verts = []
        colors = []
        for d in data:
            v =  [(d[0], cats[d[2]]-.075),
                  (d[0], cats[d[2]]+.075),
                  (d[1], cats[d[2]]+.075),
                  (d[1], cats[d[2]]-.075),
                  (d[0], cats[d[2]]-.075)]
            verts.append(v)
            colors.append(colormapping[d[2]])
        
        bars = PolyCollection(verts, facecolors=colors)
        
        x_ticks_set = []
        x_ticks_labels = []
        for j in range( int(cf.GLOBAL_VIDEO_LENGTH / cf.INTERVAL_LENGTH) - 1):
            if j % int(3600/cf.INTERVAL_LENGTH) == 0:
                start_time, end_time, startframe, endframe = _get_time(j)
                hh, mm, ss = start_time.split(":")
                hh = str(int(hh)+1).zfill(2)
                mm = ':00'
                
                x_ticks_labels.append(hh + mm)
                x_ticks_set.append(j)
            
        
        
        
        fig = plt.figure(figsize=(13,4))
        ax = fig.add_subplot(111)
        
        fig.autofmt_xdate(rotation=60)
        fig.suptitle(title, fontsize=16)
        
        ax.add_collection(bars)
        ax.autoscale()
        
        ax.set_xticks( x_ticks_set )
        ax.set_xticklabels( x_ticks_labels )
        ax.set_yticks([0.5, 1, 1.5, 2, 2.5, 3])
        behav_names.append(" ")
        ax.set_yticklabels(behav_names)
        
      
        plt.savefig(save_path + extension + '.png')
        plt.close()
    
    def _mark_truncated_images_single(single_frame_dist, 
                                      logfile,
                                      truncation_rules = truncation_rules,
                                      truncation_areas = truncation_areas,
                                      position_files = position_files,
                                      interval_length = cf.INTERVAL_LENGTH,
                                      trunc_top_standard = cf.TRUNCATION_TOP_STANDARD,
                                      trunc_bot_standard = cf.TRUNCATION_BOT_STANDARD,
                                      trunc_left_standard = cf.TRUNCATION_LEFT_STANDARD,
                                      trunc_right_standard = cf.TRUNCATION_RIGHT_STANDARD):
        
        
        truncation_upper = truncation_rules['up']
        truncation_lower = truncation_rules['bot']
        truncation_left = truncation_rules['left']
        truncation_right = truncation_rules['right']
        
        if [truncation_upper, truncation_lower, truncation_left, truncation_right] == [trunc_top_standard, trunc_bot_standard, trunc_left_standard, trunc_right_standard] and len(truncation_areas) == 0:
            return single_frame_dist
        ret = []
        for j in range(len(single_frame_dist)):   
            if np.argmax(single_frame_dist[j]) == 3:
                ret.append(single_frame_dist[j])
                continue
            
            time_interval = str(((j+1) // interval_length ) + 1).zfill(7)
            pos_info = time_interval + '.txt'
            if not os.path.exists(position_files + pos_info):
                ret.append(single_frame_dist[j])
                continue

            info_file = open(position_files + pos_info, 'r')
            info_content = info_file.read().split('\n')
            
            found_img = False
            index = 0
            
            for box_info in info_content:
                if box_info.startswith( str(j+1).zfill(7) ):
                    found_img = True
                    break
                index += 1

            if found_img:
                coordinates = info_content[index].split('-')[1].split('*')
                y1, x1, y2, x2 = int(coordinates[0]), int(coordinates[1]), int(coordinates[2]), int(coordinates[3])
                is_trunc_area = False
                
                if x2 < truncation_left or x1 > truncation_right or y1 > truncation_lower or y2 < truncation_upper:
                    ret.append( [0.0, 0.0, 0.0, 0.0, 1.0] )
                    is_trunc_area = True
                
                
                if is_trunc_area:
                    continue
                

                if len(truncation_areas) == 0:
                    ret.append( single_frame_dist[j]  )
                    continue
                    # np.array( [ [x0, y0], ..., [xn, yn] ] )
                truncation_areas = truncation_areas.reshape((-1,1,2)).astype(np.int32)

                if cv2.pointPolygonTest(truncation_areas, (x1,y1), False) >= 0 and cv2.pointPolygonTest(truncation_areas, (x1,y2), False) >= 0 and cv2.pointPolygonTest(truncation_areas, (x2,y1), False) >= 0 and cv2.pointPolygonTest(truncation_areas, (x2,y2), False) >= 0:
                    is_trunc_area = True

                if is_trunc_area:
                    ret.append( [0.0, 0.0, 0.0, 0.0, 1.0] )           
                else:
                    ret.append( single_frame_dist[j]  )
            else:
                ret.append( single_frame_dist[j]  )

        return ret
    
    def _mark_truncated_images_joint(joint_frame_dist, 
                                     truncation_rules = truncation_rules, 
                                     truncation_areas = truncation_areas,
                                     position_files = position_files,
                                     interval_len = cf.INTERVAL_LENGTH,
                                     images_per_interval = cf.IMAGES_PER_INTERVAL,
                                     behavior_names = cf.BEHAVIOR_NAMES,
                                     trunc_top_standard = cf.TRUNCATION_TOP_STANDARD,
                                      trunc_bot_standard = cf.TRUNCATION_BOT_STANDARD,
                                      trunc_left_standard = cf.TRUNCATION_LEFT_STANDARD,
                                      trunc_right_standard = cf.TRUNCATION_RIGHT_STANDARD):
        
        truncation_upper = truncation_rules['up']
        truncation_lower = truncation_rules['bot']
        truncation_left = truncation_rules['left']
        truncation_right = truncation_rules['right']
        
        if [truncation_upper, truncation_lower, truncation_left, truncation_right] == [trunc_top_standard, trunc_bot_standard, trunc_left_standard, trunc_right_standard] and len(truncation_areas) == 0:
            return joint_frame_dist
        ret = []
        for j in range(len(joint_frame_dist)):            
            if np.argmax(joint_frame_dist[j]) == 3:
                ret.append(joint_frame_dist[j])
                continue
            
            time_interval = str( j + 1 ).zfill(7)
            pos_info = time_interval + '.txt'
            if not os.path.exists(position_files + pos_info):
                ret.append(single_frame_dist[j])
                continue

            info_file = open(position_files + pos_info, 'r')
            info_content = info_file.read().split('\n')

            amount_truncated = 0
            
            for index in range(len(info_content)):
                if len(info_content[index].split('-')) < 2:
                    continue
                coordinates = info_content[index].split('-')[1].split('*')
                y1, x1, y2, x2 = int(coordinates[0]), int(coordinates[1]), int(coordinates[2]), int(coordinates[3])
                
                
                if x2 < truncation_left or x1 > truncation_right or y1 > truncation_lower or y2 < truncation_upper:
                    amount_truncated += 1
                else:    
                    area = truncation_areas
                    is_trunc_area = False
                    if len(area) > 0:
                        if cv2.pointPolygonTest(area, (x1,y1), False) >= 0 and cv2.pointPolygonTest(area, (x1,y2), False) >= 0 and cv2.pointPolygonTest(area, (x2,y1), False) >= 0 and cv2.pointPolygonTest(area, (x2,y2), False) >= 0:
                            is_trunc_area = True
                    if is_trunc_area:
                        amount_truncated += 1
                
            
            curr_dist = joint_frame_dist[j]
            
            for i in range(len(behavior_names)):
                if i == 4:
                    curr_dist[i] = amount_truncated / images_per_interval * 1.0
                else:
                    curr_dist[i] *= (1 - amount_truncated / images_per_interval) * 1.0

            curr_dist = curr_dist / np.linalg.norm(curr_dist, ord = 1)
            ret.append( curr_dist )
        #print(len(ret))
        return ret
            
    
    def _remove_out_fluctuation(phase_list, pp_rules = post_processing_rules, interval_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Removes those fluctuations between out and lying or standing if out is not too long and the rest is sufficiently long
        """
        
        def _get_current_run_endindex_out(start_index, phase_list, max_out):
            actual_behavior = phase_list[start_index-1][1]
            is_run = True
            j = start_index
            total_dur_actual_behavior = phase_list[start_index-1][0]
            total_dur_out = 0
            while is_run and j < len(phase_list):
                curr_behav = phase_list[j][1]
                if curr_behav == actual_behavior:
                    total_dur_actual_behavior += phase_list[j][0]
                    j += 1
                elif curr_behav == 3 and phase_list[j][0] <= max_out:
                    total_dur_out += phase_list[j][0]
                    j += 1
                else:
                    is_run = False
            
            return j, total_dur_actual_behavior, total_dur_out
        
        def _get_current_run_endindex_behav(start_index, phase_list, max_behav):
            actual_behavior = phase_list[start_index-1][1]
            is_run = True
            j = start_index
            total_dur_actual_behavior = phase_list[start_index-1][0]
            total_dur_out = 0
            while is_run and j < len(phase_list):
                curr_behav = phase_list[j][1]
                if curr_behav == actual_behavior and phase_list[j][0] <= max_behav:
                    total_dur_actual_behavior += phase_list[j][0]
                    j += 1
                elif curr_behav == 3:
                    total_dur_out += phase_list[j][0]
                    j += 1
                else:
                    is_run = False
            
            return j, total_dur_actual_behavior, total_dur_out
        
                    
            
        max_out = post_processing_rules['OUT_FLUCTUATION_REMOVAL_MAX']*interval_len
        max_behav = post_processing_rules['OUT_FLUCTUATION_REMOVAL_BEHAV_MAX']*interval_len
        min_actual_behavior = post_processing_rules['OUT_FLUCTUATION_REMOVAL_PERC']
        
        max_out_active = post_processing_rules['OUT_FLUCTUATION_REMOVAL_MAX_A']*interval_len
        max_behav_active = post_processing_rules['OUT_FLUCTUATION_REMOVAL_BEHAV_MAX_A']*interval_len
        min_actual_behavior_active = post_processing_rules['OUT_FLUCTUATION_REMOVAL_PERC_A']
        

        j = 1
        changes_done = False
        while j < len(phase_list)-1:
            last_behavior = phase_list[j-1][1]
            current_behavior = phase_list[j][1]
            next_behavior = phase_list[j+1][1]

            if current_behavior == 3:
                
                if last_behavior == 0:
                    out_thresh = max_out_active
                    behavior_thresh_dur = max_behav_active
                    behavior_thresh_perc = min_actual_behavior_active                
                else:
                    out_thresh = max_out
                    behavior_thresh_dur = max_behav
                    behavior_thresh_perc = min_actual_behavior                    
                
                if phase_list[j][0] > out_thresh:
                    j += 1
                    continue
                if next_behavior != last_behavior:
                    j += 1
                    continue
                end_index_out, dur_behav_in_out, dur_out_in_out = _get_current_run_endindex_out(j, phase_list, out_thresh)
                end_index_behav, dur_behav_in_behav, dur_out_in_behav = _get_current_run_endindex_behav(j, phase_list, behavior_thresh_dur)
                
                
                if dur_out_in_out / (dur_out_in_out + dur_behav_in_out) <= behavior_thresh_perc:
                    changes_done = True
                    for h in range(j, end_index_out):
                        phase_list[h][1] = last_behavior
                    j = end_index_out + 1
                elif dur_behav_in_behav / (dur_out_in_behav + dur_behav_in_behav) <= behavior_thresh_perc and j >= end_index_behav + 2:
                    changes_done = True
                    for h in range(j, end_index_behav):
                        phase_list[h][1] = 3
                    j = end_index_behav + 1
                else:
                    j += 1
                    continue
                
                
            else:
                j += 1
        
        phase_list = _remove_truncated_images(phase_list) 
        phase_list = _join_phases(phase_list)
        

        return phase_list, changes_done
                    
                
        
        
    
    def _time_shift(dist, length, int_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Shifts any behavioral sequence by the rolling average order
        """
        if len(dist) <= 1 or length <= 0:
            return dist
        
        dist[0][0] -= length*int_len
        dist[0][3] -= length
        
        for j in range(1, len(dist)):
            dist[j][2] -= length
            dist[j][3] -= length
        
        dist[-1][0] += length*int_len
        dist[-1][3] += length
        
        return dist
    
    def _get_video_length(start, end):
            if end > start:
                return end-start
            
            return 24 + (end-start)
        
    def _add_observation_time_sf(single_frame_dist, vid_start = current_video_start, 
                          vid_end = current_video_end,
                          observation_start = cf.GLOBAL_STARTING_TIME,
                          observation_end = cf.GLOBAL_ENDING_TIME):
        
        
        
        observation_intervals = _get_video_length(observation_start, observation_end)*3600 
        
        ret = []
        frames_fill_start = (vid_start-observation_start)*60*60
        frames_fill_end = (observation_end - vid_end)*60*60
        
        
        if frames_fill_start > 0:
            for i in range(frames_fill_start):
                ret.append( [0.0, 0.0, 0.0, 1.0, 0.0] )
        elif frames_fill_start < 0:
            single_frame_dist = single_frame_dist[(-1)*frames_fill_start:]
            
        for j in range(len(single_frame_dist)):
            ret.append( single_frame_dist[j] )
        if frames_fill_end > 0:
            for i in range(frames_fill_end):
                ret.append( [0.0, 0.0, 0.0, 1.0, 0.0] )
        elif frames_fill_end < 0:
            ret = ret[:frames_fill_end]
        
        missing_intervals = observation_intervals - len(ret)
        
        if missing_intervals > 0:
            for j in range(missing_intervals):
                ret.append( [0.0, 0.0, 0.0, 1.0, 0.0] )

        return ret

    
    def _add_observation_time_mf(multiple_frame_dist, vid_start = current_video_start, 
                          vid_end = current_video_end,
                          observation_start = cf.GLOBAL_STARTING_TIME,
                          observation_end = cf.GLOBAL_ENDING_TIME,
                          interval = cf.INTERVAL_LENGTH):
        
        
        observation_intervals = _get_video_length(observation_start, observation_end)*3600 // interval 
        intervals_per_hour = 3600//interval
        
        ret = []
        frames_fill_start = (vid_start-observation_start) * intervals_per_hour
        frames_fill_end = (observation_end - vid_end) * intervals_per_hour
        
                
        if frames_fill_start > 0:
            for i in range(frames_fill_start):
                ret.append( [0.0, 0.0, 0.0, 1.0, 0.0] )
        elif frames_fill_start < 0:
            # rounding issue TODO: make more smooth!
            round_issue = 0
            if vid_start-observation_start == -2:
                round_issue = 2
            elif vid_start-observation_start == -1:
                round_issue = 3
            multiple_frame_dist = multiple_frame_dist[(-1)*frames_fill_start-round_issue:]
            
        for j in range(len(multiple_frame_dist)):
            ret.append( multiple_frame_dist[j] )
        
        if frames_fill_end > 0:
            for i in range(frames_fill_end+1):
                ret.append( [0.0, 0.0, 0.0, 1.0, 0.0] )
        elif frames_fill_end < 0:
            # rounding issue TODO: make more smooth!
            if observation_end - vid_end == -2:
                round_issue = 2
            elif observation_end - vid_end == -1:
                round_issue = 1
            ret = ret[:frames_fill_end-round_issue]
            
        missing_intervals = observation_intervals - len(ret)
        if missing_intervals > 0:
            for j in range(missing_intervals):
                ret.append( [0.0, 0.0, 0.0, 1.0, 0.0] )
        return ret
    
    
    
    
    def _add_out_regulations_start(ensemble_sparse, out_duration):
        ensemble_sparse[:out_duration] = [3]*out_duration
        return ensemble_sparse
    
    
    # initialise logfile
    ensure_dir(output_folder_prediction)
    with open(logfile, 'a+') as file:
        file.writelines(['***** {} {}\n'.format(individual_code, datum), '+++ truncation\n'])
    
    # read single frame file and merge it to time intervals
    single_frame_dist = _read_csv(single_frame_csv)
    single_frame_dist = _mark_truncated_images_single(single_frame_dist, logfile)
    #print( [(j, j//7, single_frame_dist[j][4]) for j in range(len(single_frame_dist)) if single_frame_dist[j][4] > 0] )
    single_frame_dist = _add_observation_time_sf(single_frame_dist)   
    single_frame_dist = _apply_rolling_average( single_frame_dist, 
                                               post_processing_rules['ROLLING_AVERAGE_SINGLE_FRAMES'], 
                                               post_processing_rules['ROLLING_AVERAGE_WEIGHTS'],
                                               post_processing_rules['ROLLING_AVERAGE_POTENCY_SF']
                                               )
    sf_cumulated = _single_frame_prediction_to_intervalprediction(single_frame_dist)
    
    
    # read time interval file and apply rolling average
    joint_image_dist = _read_csv(joint_interval_csv)
    joint_image_dist = _mark_truncated_images_joint(joint_image_dist)
    #print( [(j, round(joint_image_dist[j][4], 2)) for j in range(len(joint_image_dist)) if joint_image_dist[j][4] > 0] )
    joint_image_dist = _add_observation_time_mf(joint_image_dist)    
    joint_image_dist = _apply_rolling_average(joint_image_dist, 
                                              post_processing_rules['ROLLING_AVERAGE_JOINT_IMAGES'], 
                                              post_processing_rules['ROLLING_AVERAGE_WEIGHTS'],
                                              post_processing_rules['ROLLING_AVERAGE_POTENCY_MF']
                                              )
    # merge the two prediction types
    ensemble_dist = _calculate_joint_prediction( dist1 = joint_image_dist, 
                                                dist2 = sf_cumulated, 
                                                weights = post_processing_rules['WEIGHTS_NETWORKS'])
        
    ensemble_dist = _apply_rolling_average(ensemble_dist, 
                                           post_processing_rules['ROLLING_AVERAGE_ENSEMBLE'], 
                                           post_processing_rules['ROLLING_AVERAGE_WEIGHTS'],
                                           post_processing_rules['ROLLING_AVERAGE_POTENCY_ENSEMBLE'])
    
    
    
    save_path_csv = output_folder_prediction + 'raw_csv/ensemble/'
    filename_csv = datum + '_' + individual_code + '_ensemble.csv'
    _write_prediction_csv(ensemble_dist, save_path_csv, filename_csv)
    
    with open(logfile, 'a+') as file:
        file.writelines(['+++ out (seconds new_behavior start_interval) \n'])
        
    # apply post processing rules
    ensemble_sparse = _sparse_encoding(ensemble_dist)
    
    if individual_code in out_regulations.keys():
        if datum in out_regulations[individual_code][0]:
            out_dur = out_regulations[individual_code][1]
            ensemble_sparse = _add_out_regulations_start(ensemble_sparse, out_dur)
            
    ensemble_sparse = _extract_single_phases( ensemble_sparse )
    ensemble_sparse = _remove_truncated_images(ensemble_sparse)    
    ensemble_sparse = _remove_out(ensemble_sparse, logfile)    
    ensemble_sparse = _map_behaviors(ensemble_sparse)
    
    
    
    x = True
    c_count = 0
    while x:
        c_count += 1
        if c_count > 200: # sanity check whether no more changes are done
            ens_copy = copy.deepcopy(ensemble_sparse)
            ensemble_sparse, x = _remove_out_fluctuation(ensemble_sparse)
            if ens_copy == ensemble_sparse:
                x = False
        else:
            ensemble_sparse, x = _remove_out_fluctuation(ensemble_sparse)   
    
    ensemble_sparse = _remove_short_phases(ensemble_sparse)   
    ensemble_sparse = _time_shift(ensemble_sparse, length = post_processing_rules['ROLLING_AVERAGE_ENSEMBLE'] - 1)
    

    
    phases_ordered = [ [x[0] for x in ensemble_sparse if x[1] == 0], 
                      [x[0] for x in ensemble_sparse if x[1] == 1],
                      [x[0] for x in ensemble_sparse if x[1] == 2],
                      [x[0] for x in ensemble_sparse if x[1] == 3],
                      [x[0] for x in ensemble_sparse if x[1] == 4]
                     ]
    sparse_postprocessed = _get_sparse_intervals_from_phases(ensemble_sparse)
    
    outputfolder = output_folder_prediction + 'final/'
    filename = datum + '_' + individual_code + '_statistics' 
    _write_xlsx_statistics(phases_ordered, ensemble_sparse, sparse_postprocessed, outputfolder, filename)
    
    save_path = outputfolder + datum + '_' + individual_code + '_timeline'
    title = datum + '_' + individual_code
    _draw_timeline(ensemble_sparse, save_path, title)